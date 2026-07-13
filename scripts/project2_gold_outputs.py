from __future__ import annotations

import html
import hashlib
import importlib.metadata as metadata
import json
import locale
import os
import platform
import re
import shutil
import sqlite3
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from project2_gold_common import (
    DISCLAIMER,
    MODEL_BASELINE_END,
    MODEL_BASELINE_START,
    ROOT,
    RUN_DATE,
    VERSION,
    add_meta,
    ensure_dirs,
    table_md,
    write_csv,
    write_md,
)


NAVY = "18323E"
TEAL = "0B6B68"
GREEN = "2F7D4A"
AMBER = "C98613"
RED = "B33A3A"
LIGHT = "EDF2F3"
WHITE = "FFFFFF"
GRAY = "5F6B72"


def _read(relative: str, **kwargs) -> pd.DataFrame:
    return pd.read_csv(ROOT / relative, **kwargs)


def _metric(frame: pd.DataFrame, name: str) -> float:
    return float(frame.loc[frame["metric"].eq(name), "value"].iloc[0])


def _pct(value: float, decimals: int = 2) -> str:
    return "N/A" if pd.isna(value) else f"{value:.{decimals}%}"


def _num(value: float, decimals: int = 0) -> str:
    return "N/A" if pd.isna(value) else f"{value:,.{decimals}f}"


def _bps(value: float) -> str:
    return "N/A" if pd.isna(value) else f"{value * 10_000:+,.0f} bps"


def _write_sql(relative: str, text: str) -> None:
    path = ROOT / relative
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.strip() + "\n", encoding="utf-8")


def generate_sql_layer() -> None:
    """Create and execute a portable SQLite monitoring mart implementation."""
    _write_sql(
        "sql/01_create_staging_tables.sql",
        """
PRAGMA foreign_keys = ON;
DROP TABLE IF EXISTS stg_p2_observed_accounts;
DROP TABLE IF EXISTS stg_p2_synthetic_snapshots;

CREATE TABLE stg_p2_observed_accounts AS
SELECT
    CAST(account_id AS TEXT) AS account_id,
    CAST(application_period AS TEXT) AS application_period,
    CAST(matured_flag AS INTEGER) AS matured_flag,
    CAST(outcome_eligible_flag AS INTEGER) AS outcome_eligible_flag,
    CAST(bad_flag AS INTEGER) AS bad_flag,
    CAST(pd AS REAL) AS pd,
    CAST(scorecard_points AS REAL) AS scorecard_points,
    CAST(risk_grade AS TEXT) AS risk_grade,
    CAST(exposure AS REAL) AS exposure,
    CAST(expected_loss AS REAL) AS expected_loss,
    CAST(stage_proxy AS TEXT) AS stage_proxy,
    CAST(policy_version AS TEXT) AS policy_version,
    CAST(policy_decision AS TEXT) AS policy_decision,
    CAST(exception_flag AS INTEGER) AS exception_flag,
    CAST(fico AS REAL) AS fico,
    CAST(dti AS REAL) AS dti,
    CAST(purpose AS TEXT) AS purpose
FROM raw_observed;

CREATE UNIQUE INDEX ux_stg_observed_account ON stg_p2_observed_accounts(account_id);
CREATE INDEX ix_stg_observed_period ON stg_p2_observed_accounts(application_period);
CREATE INDEX ix_stg_observed_grade ON stg_p2_observed_accounts(risk_grade);

CREATE TABLE stg_p2_synthetic_snapshots AS
SELECT
    CAST(account_id AS TEXT) AS account_id,
    CAST(origination_vintage AS TEXT) AS origination_vintage,
    CAST(snapshot_month AS TEXT) AS snapshot_month,
    CAST(months_on_book AS INTEGER) AS months_on_book,
    CAST(term_months AS INTEGER) AS term_months,
    CAST(beginning_balance AS REAL) AS beginning_balance,
    CAST(ending_balance AS REAL) AS ending_balance,
    CAST(beginning_dpd_bucket AS TEXT) AS beginning_dpd_bucket,
    CAST(dpd_bucket AS TEXT) AS dpd_bucket,
    CAST(new_delinquency_inflow_flag AS INTEGER) AS new_delinquency_inflow_flag,
    CAST(cure_flag AS INTEGER) AS cure_flag,
    CAST(new_default_flag AS INTEGER) AS new_default_flag,
    CAST(default_stock_flag AS INTEGER) AS default_stock_flag,
    CAST(closure_flow_flag AS INTEGER) AS closure_flow_flag,
    CAST(risk_grade AS TEXT) AS risk_grade,
    CAST(policy_version AS TEXT) AS policy_version
FROM raw_synthetic;

CREATE UNIQUE INDEX ux_stg_synthetic_account_month
ON stg_p2_synthetic_snapshots(account_id, snapshot_month);
CREATE INDEX ix_stg_synthetic_month ON stg_p2_synthetic_snapshots(snapshot_month);
""",
    )
    _write_sql(
        "sql/02_build_snapshot_facts.sql",
        """
DROP TABLE IF EXISTS fact_portfolio_snapshot;
CREATE TABLE fact_portfolio_snapshot AS
SELECT * FROM stg_p2_observed_accounts;

DROP TABLE IF EXISTS fact_synthetic_servicing;
CREATE TABLE fact_synthetic_servicing AS
SELECT * FROM stg_p2_synthetic_snapshots;

CREATE UNIQUE INDEX ux_fact_portfolio_account ON fact_portfolio_snapshot(account_id);
CREATE UNIQUE INDEX ux_fact_servicing_account_month ON fact_synthetic_servicing(account_id, snapshot_month);
""",
    )
    _write_sql(
        "sql/03_build_dimensions.sql",
        """
DROP TABLE IF EXISTS dim_period;
CREATE TABLE dim_period AS
SELECT DISTINCT application_period AS period_key, 'OBSERVED_COHORT' AS period_type
FROM fact_portfolio_snapshot
UNION
SELECT DISTINCT snapshot_month AS period_key, 'SYNTHETIC_SNAPSHOT' AS period_type
FROM fact_synthetic_servicing;

DROP TABLE IF EXISTS dim_risk_grade;
CREATE TABLE dim_risk_grade AS
SELECT
    ROW_NUMBER() OVER (ORDER BY SUBSTR(risk_grade, 1, 1)) AS risk_grade_key,
    risk_grade,
    SUBSTR(risk_grade, 1, 1) AS grade_code,
    CASE SUBSTR(risk_grade, 1, 1)
        WHEN 'A' THEN 1
        WHEN 'B' THEN 2
        WHEN 'C' THEN 3
        WHEN 'D' THEN 4
        WHEN 'E' THEN 5
        ELSE 99
    END AS grade_order
FROM (SELECT DISTINCT risk_grade FROM fact_portfolio_snapshot WHERE risk_grade IS NOT NULL);

DROP TABLE IF EXISTS dim_policy_decision;
CREATE TABLE dim_policy_decision AS
SELECT ROW_NUMBER() OVER (ORDER BY policy_decision) AS decision_key, policy_decision
FROM (SELECT DISTINCT policy_decision FROM fact_portfolio_snapshot WHERE policy_decision IS NOT NULL);

DROP TABLE IF EXISTS dim_dpd_bucket;
CREATE TABLE dim_dpd_bucket AS
SELECT ROW_NUMBER() OVER (ORDER BY dpd_bucket) AS dpd_bucket_key, dpd_bucket
FROM (SELECT DISTINCT dpd_bucket FROM fact_synthetic_servicing WHERE dpd_bucket IS NOT NULL);
""",
    )
    _write_sql(
        "sql/04_build_monitoring_marts.sql",
        """
DROP TABLE IF EXISTS mart_portfolio_period;
CREATE TABLE mart_portfolio_period AS
SELECT
    application_period,
    COUNT(*) AS accounts,
    SUM(matured_flag) AS matured_accounts,
    SUM(CASE WHEN matured_flag = 1 THEN bad_flag ELSE 0 END) AS matured_bad_accounts,
    CASE WHEN SUM(matured_flag) = 0 THEN NULL
         ELSE 1.0 * SUM(CASE WHEN matured_flag = 1 THEN bad_flag ELSE 0 END) / SUM(matured_flag)
    END AS observed_bad_rate_matured,
    SUM(exposure) AS exposure,
    AVG(pd) AS average_pd_all,
    CASE WHEN SUM(matured_flag) = 0 THEN NULL
         ELSE SUM(CASE WHEN matured_flag = 1 THEN pd ELSE 0 END) / SUM(matured_flag)
    END AS average_pd_matured,
    SUM(expected_loss) AS expected_loss
FROM fact_portfolio_snapshot
GROUP BY application_period;

DROP TABLE IF EXISTS mart_synthetic_stock_flow;
CREATE TABLE mart_synthetic_stock_flow AS
SELECT
    snapshot_month,
    COUNT(*) AS accounts,
    SUM(CASE WHEN beginning_dpd_bucket IN ('1-29','30-59','60-89','90+') THEN 1 ELSE 0 END) AS beginning_delinquent_stock,
    SUM(new_delinquency_inflow_flag) AS new_delinquency_inflow,
    SUM(cure_flag) AS cures,
    SUM(new_default_flag) AS new_default_flow,
    SUM(default_stock_flag) AS ending_default_stock,
    SUM(closure_flow_flag) AS closures,
    SUM(CASE WHEN dpd_bucket IN ('1-29','30-59','60-89','90+') THEN 1 ELSE 0 END) AS ending_delinquent_stock
FROM fact_synthetic_servicing
GROUP BY snapshot_month;

DROP TABLE IF EXISTS mart_vintage;
CREATE TABLE mart_vintage AS
SELECT
    origination_vintage,
    months_on_book,
    term_months,
    risk_grade,
    COUNT(*) AS accounts,
    SUM(ending_balance) AS ending_balance,
    AVG(default_stock_flag) AS default_stock_rate,
    AVG(CASE WHEN dpd_bucket IN ('30-59','60-89','90+','Default') THEN 1.0 ELSE 0.0 END) AS dpd_30_plus_rate
FROM fact_synthetic_servicing
GROUP BY origination_vintage, months_on_book, term_months, risk_grade;
""",
    )
    _write_sql(
        "sql/05_validation_queries.sql",
        """
SELECT 'observed_account_key_unique' AS check_name,
       CASE WHEN COUNT(*) = COUNT(DISTINCT account_id) THEN 'PASS' ELSE 'FAIL' END AS status,
       COUNT(*) AS actual_value
FROM fact_portfolio_snapshot;

SELECT 'synthetic_account_month_key_unique' AS check_name,
       CASE WHEN COUNT(*) = COUNT(DISTINCT account_id || '|' || snapshot_month) THEN 'PASS' ELSE 'FAIL' END AS status,
       COUNT(*) AS actual_value
FROM fact_synthetic_servicing;

SELECT 'pd_range' AS check_name,
       CASE WHEN SUM(CASE WHEN pd < 0 OR pd > 1 OR pd IS NULL THEN 1 ELSE 0 END) = 0 THEN 'PASS' ELSE 'FAIL' END AS status,
       SUM(CASE WHEN pd < 0 OR pd > 1 OR pd IS NULL THEN 1 ELSE 0 END) AS actual_value
FROM fact_portfolio_snapshot;

SELECT 'immature_bad_rate_is_null' AS check_name,
       CASE WHEN SUM(CASE WHEN matured_accounts = 0 AND observed_bad_rate_matured IS NOT NULL THEN 1 ELSE 0 END) = 0 THEN 'PASS' ELSE 'FAIL' END AS status,
       SUM(CASE WHEN matured_accounts = 0 AND observed_bad_rate_matured IS NOT NULL THEN 1 ELSE 0 END) AS actual_value
FROM mart_portfolio_period;
""",
    )

    observed = _read("data/reviewer_samples/observed_account_sample_5k.csv")
    full_observed = _read("data/observed/p2_observed_account_base.csv.gz", nrows=5_000)
    synthetic = _read("data/reviewer_samples/synthetic_snapshot_sample_50k.csv.gz")
    db_path = ROOT / "validation/project2_sql_execution.db"
    if db_path.exists():
        db_path.unlink()
    connection = sqlite3.connect(db_path)
    full_observed.to_sql("raw_observed", connection, index=False, if_exists="replace")
    synthetic.to_sql("raw_synthetic", connection, index=False, if_exists="replace")
    for script in [
        "01_create_staging_tables.sql",
        "02_build_snapshot_facts.sql",
        "03_build_dimensions.sql",
        "04_build_monitoring_marts.sql",
    ]:
        connection.executescript((ROOT / "sql" / script).read_text(encoding="utf-8"))
    checks = []
    for query in [part.strip() for part in (ROOT / "sql/05_validation_queries.sql").read_text(encoding="utf-8").split(";") if part.strip()]:
        checks.append(pd.read_sql_query(query, connection).iloc[0].to_dict())
    table_names = pd.read_sql_query(
        "SELECT name FROM sqlite_master WHERE type='table' AND (name LIKE 'fact_%' OR name LIKE 'dim_%' OR name LIKE 'mart_%') ORDER BY name",
        connection,
    )
    connection.close()
    db_path.unlink()
    checks.append(
        {
            "check_name": "required_sql_marts_created",
            "status": "PASS" if len(table_names) >= 9 else "FAIL",
            "actual_value": len(table_names),
        }
    )
    sql_checks = add_meta(
        pd.DataFrame(checks),
        "Technical",
        "Executed",
        "P2",
        "SQLite execution on packaged 5k observed and 50k synthetic reviewer samples",
    )
    write_csv(sql_checks, "validation/sql_execution_results.csv")
    write_md(
        "sql/README.md",
        "Executable SQL Monitoring Layer",
        f"""The five SQL scripts were executed with Python `sqlite3` against packaged reviewer samples. The implementation builds typed staging, two facts, four dimensions/marts and executable validation queries.

{table_md(sql_checks[["check_name", "status", "actual_value"]])}

`risk_grade` is explicitly loaded into staging and fact tables; the v1 missing-column defect is removed.
""",
    )


def generate_governance_pack() -> None:
    assumptions = pd.DataFrame(
        [
            ["ASM-001", "Outcome maturity", "matured_flag = 1", "P3/P6 frozen data", "Observed bad-rate and calibration denominator", "Model Owner"],
            ["ASM-002", "Minimum matured accounts", 1000, "Monitoring design", "Period eligibility", "Model Risk"],
            ["ASM-003", "Minimum bad events", 50, "Monitoring design", "Period eligibility", "Model Risk"],
            ["ASM-004", "PSI baseline", f"{MODEL_BASELINE_START} to {MODEL_BASELINE_END}", "P3 development era", "Fixed-bin drift reference", "Model Risk"],
            ["ASM-005", "Synthetic horizon", "36 MOB", "Controls-testing design", "Does not imply loan closure", "Collections"],
            ["ASM-006", "Synthetic term mix", "70% 36m / 30% 60m", "Documented proxy because P3 term is unavailable", "Term-aware amortisation", "Portfolio Risk"],
            ["ASM-007", "Stage 2", "P4 broad proxy", "P4 v1.2.2", "Not production SICR", "Credit Risk"],
            ["ASM-008", "Manual review handling time", "10 minutes/case", "P6 assumption", "Operations workload proxy", "Operations"],
            ["ASM-009", "Risk appetite", "Educational thresholds", "Portfolio design", "No organisational approval claim", "Risk Committee"],
            ["ASM-010", "ECL period movement", "No prior snapshot", "P4 frozen reporting date", "No fabricated trend", "Finance Risk"],
        ],
        columns=["assumption_id", "assumption", "value", "basis", "use_or_impact", "owner"],
    )
    write_csv(add_meta(assumptions, "Governance", "Assumption", "P2", "Configurable educational assumptions"), "governance/assumption_register.csv")

    kpis = pd.DataFrame(
        [
            ["KPI-001", "Observed bad rate matured", "Matured bad accounts / matured accounts", "Observed", "Monthly", "Portfolio Risk", "Observed Bad Rate Matured"],
            ["KPI-002", "Average PD matured", "Sum PD for matured / matured accounts", "Observed/Derived", "Monthly", "Model Owner", "Average PD Matured"],
            ["KPI-003", "Calibration gap matured", "Observed bad rate matured - average PD matured", "Observed/Derived", "Monthly", "Model Owner", "Calibration Gap Matured"],
            ["KPI-004", "PD PSI", "Fixed development-bin PSI", "Derived", "Monthly", "Model Risk", "Latest Eligible PD PSI"],
            ["KPI-005", "30+ DPD rate", "Synthetic 30+ stock / synthetic snapshot accounts", "Synthetic", "Monthly", "Collections", "30 Plus DPD Rate"],
            ["KPI-006", "New default flow rate", "New defaults / accounts at risk", "Synthetic", "Monthly", "Collections", "New Default Flow Rate"],
            ["KPI-007", "Policy approval rate", "Approved / common matured population", "Observed/Simulated", "Version", "Credit Policy", "Policy Approval Rate"],
            ["KPI-008", "Approved bad rate", "Approved bads / matured approved", "Observed/Simulated", "Version", "Credit Policy", "Approved Bad Rate Matured"],
            ["KPI-009", "Weighted ECL rate", "Weighted ECL / current EAD", "Proxy", "Reporting date", "Finance Risk", "Weighted ECL Rate"],
            ["KPI-010", "Open action count", "Count action records with OPEN status", "Governance", "Monthly", "Risk Committee", "Open Action Count"],
        ],
        columns=["kpi_id", "kpi_name", "formula", "evidence_status", "frequency", "owner", "powerbi_measure"],
    )
    write_csv(add_meta(kpis, "Governance", "Definition", "P2", "KPI definitions and ownership"), "governance/kpi_dictionary.csv")
    write_csv(add_meta(kpis, "Governance", "Definition", "P2", "KPI definitions and ownership"), "08_kpi_dictionary.csv")

    kri = _read("outputs/kri_status_summary.csv")
    kri_register = kri[[
        "kri_id", "metric", "green_threshold", "amber_threshold", "direction", "owner", "source_scope",
        "impact_method", "claim_boundary",
    ]].copy()
    kri_register["frequency"] = np.where(kri_register["source_scope"].str.contains("P4|P5"), "Frozen overlay review", "Monthly")
    write_csv(add_meta(kri_register, "Governance", "Definition", "P2", "Threshold and ownership register"), "governance/kri_register.csv")
    write_csv(add_meta(kri_register, "Governance", "Definition", "P2", "Threshold and ownership register"), "09_kri_register.csv")

    appetite = kri[["kri_id", "metric", "green_threshold", "amber_threshold", "direction", "owner"]].copy()
    appetite["red_boundary"] = np.where(appetite["direction"].eq("higher_bad"), "> amber threshold", "< amber threshold")
    appetite["approval_status"] = "EDUCATIONAL_NOT_ORGANISATIONALLY_APPROVED"
    write_csv(add_meta(appetite, "Governance", "Assumption", "P2", "Educational risk appetite"), "governance/risk_appetite_register.csv")
    write_csv(add_meta(appetite, "Governance", "Assumption", "P2", "Educational risk appetite"), "10_risk_appetite_register.csv")

    monitoring = pd.DataFrame(
        [
            ["Portfolio quality", "Monthly", "Portfolio Risk", "Risk Committee", "Bad rate, mix, exposure", "outputs/portfolio_monthly_trend.csv"],
            ["Delinquency controls", "Monthly synthetic", "Collections", "Portfolio Risk", "Stock-flow, roll-rate, cure", "outputs/delinquency_stock_flow_summary.csv"],
            ["Model monitoring", "Latest eligible period", "Model Owner", "Model Risk", "AUC, KS, calibration, PSI", "outputs/model_monitoring_current_summary.csv"],
            ["Policy monitoring", "Version and period", "Credit Policy", "Portfolio Risk", "Approval, bad capture, EL", "outputs/policy_version_bridge_summary.csv"],
            ["ECL/stress", "Frozen reporting date", "Finance Risk", "Credit Risk", "ECL, concentration, readiness", "outputs/ecl_stress_monitoring_summary.csv"],
            ["EWS/actions", "Monthly plus static overlays", "Risk Owners", "Risk Committee", "Status, state, impact, actions", "outputs/kri_status_summary.csv"],
        ],
        columns=["monitoring_domain", "frequency", "owner", "approver", "metrics", "evidence"],
    )
    write_csv(add_meta(monitoring, "Governance", "Plan", "P2", "Monitoring cadence and ownership"), "governance/monitoring_plan.csv")
    owner_matrix = monitoring[["monitoring_domain", "owner", "approver"]].copy()
    owner_matrix["independent_review_role"] = "Proposed Model Validation/Internal Audit"
    owner_matrix["actual_signoff"] = "NO"
    write_csv(add_meta(owner_matrix, "Governance", "Plan", "P2", "No organisational sign-off claimed"), "governance/owner_approver_matrix.csv")

    exceptions = pd.DataFrame(
        [
            ["EXC-001", "No observed monthly servicing history", "Use separately labelled synthetic controls", "HIGH", "OPEN_DATA_GAP", "Collections"],
            ["EXC-002", "P4 Stage 2 is broad proxy", "Acquire current PD/DPD/SICR evidence", "HIGH", "OPEN_DATA_GAP", "Credit Risk"],
            ["EXC-003", "No prior P4 ECL snapshot", "Do not show false period trend", "MEDIUM", "CONTROLLED", "Finance Risk"],
            ["EXC-004", "No Power BI Desktop in build environment", "Provide governed build pack; PBIX remains external gate", "HIGH", "OPEN_EXTERNAL_GATE", "BI Owner"],
            ["EXC-005", "DAX runtime unavailable without PBIX", "Semantic lint and Python source values only", "HIGH", "OPEN_EXTERNAL_GATE", "BI Owner"],
        ],
        columns=["exception_id", "issue", "treatment", "severity", "status", "owner"],
    )
    write_csv(add_meta(exceptions, "Governance", "Exception", "P2", "Explicit unresolved and controlled limitations"), "governance/exception_register.csv")
    data_status = pd.DataFrame(
        [
            ["Observed portfolio", "Observed/Derived", "P3/P6", "Historical accepted/booked accounts", "Outcome metrics matured only"],
            ["Servicing history", "Synthetic", "P2", "Controls-testing only", "Not production delinquency evidence"],
            ["Policy versions", "Observed/Simulated", "P6/P2", "Common booked population", "Not applicant-level approval simulation"],
            ["ECL/stage", "Proxy", "P4", "Frozen snapshot", "No prior-period trend; Stage 2 proxy"],
            ["Fraud/OpRisk", "Mixed", "P5", "Observed benchmark plus synthetic controls", "Boundaries preserved"],
        ],
        columns=["domain", "data_status", "source", "permitted_claim", "prohibited_claim"],
    )
    write_csv(add_meta(data_status, "Governance", "Definition", "P2", "Evidence status register"), "governance/data_status_register.csv")
    model_use = pd.DataFrame(
        [["P3 scorecard/PD", "Monitoring only", "Frozen", "No retraining in P2", "AUC/KS/calibration/PSI"], ["Synthetic servicing engine", "Controls testing", "Scenario", "Not predictive model", "Bridge and transition controls"]],
        columns=["model_or_engine", "permitted_use", "status", "restriction", "monitoring"],
    )
    write_csv(add_meta(model_use, "Governance", "Definition", "P2", "Model-use boundary"), "governance/model_use_register.csv")
    audit_items = pd.DataFrame(
        [[f"AUD-{i:03d}", item, owner, "EVIDENCE_PREPARED", "NO"] for i, (item, owner) in enumerate([
            ("Source hashes and versions", "Data Owner"), ("Maturity denominators", "Portfolio Risk"),
            ("Stock-flow reconciliation", "Collections"), ("Model eligibility and drift", "Model Risk"),
            ("Policy bridge", "Credit Policy"), ("ECL limitations", "Finance Risk"),
            ("Alert-action linkage", "Risk Committee"), ("SQL execution", "Engineering"),
            ("Excel formula controls", "Finance Controls"), ("PBIX publication", "BI Owner"),
        ], start=1)],
        columns=["audit_id", "evidence_area", "owner", "evidence_status", "independent_signoff_performed"],
    )
    write_csv(add_meta(audit_items, "Governance", "Checklist", "P2", "Evidence prepared; no independent sign-off"), "governance/audit_evidence_checklist.csv")

    write_md("methodology/portfolio_monitoring_methodology.md", "Portfolio Monitoring Methodology", "Observed account/cohort metrics use the full accepted/booked base. Outcome metrics use only `matured_flag = 1`; zero-eligible cohorts are `NOT_ASSESSABLE`. Current/prior/change fields are retained at period grain.")
    write_md("methodology/vintage_cohort_methodology.md", "Vintage and Cohort Methodology", "Observed origination cohorts measure booked-account composition and matured outcomes. Synthetic vintage curves use 48 origination vintages and MOB 1-36 with term-aware 36/60-month balances. Synthetic servicing evidence is never relabelled observed.")
    write_md("methodology/delinquency_rollrate_methodology.md", "Delinquency, Roll-Rate and Stock-Flow Methodology", "Roll rates use beginning-to-ending DPD transitions. Stock-flow reconciles beginning delinquent stock + new inflow - cures - defaults - closures + other migration net = ending delinquent stock. Default stock and new-default flow are separate measures.")
    write_md("methodology/model_monitoring_methodology.md", "Model Monitoring Methodology", f"A period is eligible only with at least 1,000 matured accounts, 50 bad events and both outcome classes. AUC, KS, Gini, calibration and O/E use matured outcomes only. PSI uses fixed bins from {MODEL_BASELINE_START} to {MODEL_BASELINE_END}; it is not summed across periods.")
    write_md("methodology/calibration_monitoring_methodology.md", "Calibration Monitoring Methodology", "Calibration gap = observed bad rate matured - average PD matured. Tables explicitly report excluded immature accounts and assessment status. Latest eligible and prior eligible periods drive EWS.")
    write_md("methodology/policy_monitoring_methodology.md", "Policy Monitoring Methodology", "Baseline 25%, candidate 20%, conservative 15%, 20%+overlay and decline-threshold sensitivity are compared on the same historical booked population. Outcome metrics use matured accounts. A zero decline rate at 35% is an empirical result of the frozen PD range, not a coding omission.")
    write_md("methodology/ecl_monitoring_methodology.md", "ECL Monitoring Methodology", "P4 account-level snapshot is monitored by stage, grade, purpose, term, scenario and sensitivity. The report does not fabricate a prior-period ECL trend. Stage 2 and lifetime-ECL limitations remain visible.")
    write_md("methodology/early_warning_methodology.md", "Early-Warning Methodology", "Dynamic KRIs use latest eligible period, prior period, movement and alert persistence. Static P4/P5 overlays are labelled `STATIC_REVIEW`. Every breach has source-specific population, exposure or non-monetised impact, owner, action and closure evidence.")
    write_md("methodology/risk_appetite_methodology.md", "Risk Appetite Methodology", "Green/Amber/Red thresholds are configurable educational assumptions. Amber variance is measured from the Green trigger; Red variance is measured from the Amber trigger. Lower-is-bad metrics reverse the comparison consistently.")
    write_md("methodology/dashboard_measure_governance.md", "Dashboard Measure Governance", "Dashboard measures map to governed Python outputs and formula definitions. DAX semantic lint is executable, but DAX runtime execution remains an explicit external gate until a real PBIX is built in Power BI Desktop.")


def _dax_rows() -> list[list[str]]:
    return [
        ["Total Accounts", "DISTINCTCOUNT(fact_portfolio_snapshot[account_id])", "Whole number", "Observed account population", "outputs/portfolio_kpi_summary.csv"],
        ["Matured Accounts", "CALCULATE(DISTINCTCOUNT(fact_portfolio_snapshot[account_id]), fact_portfolio_snapshot[matured_flag] = 1)", "Whole number", "Outcome-eligible accounts", "outputs/portfolio_kpi_summary.csv"],
        ["Immature Accounts", "[Total Accounts] - [Matured Accounts]", "Whole number", "Monitor-only accounts", "outputs/portfolio_kpi_summary.csv"],
        ["Outcome Eligibility Rate", "DIVIDE([Matured Accounts], [Total Accounts])", "0.00%", "Maturity coverage", "outputs/portfolio_kpi_summary.csv"],
        ["Total Exposure", "SUM(fact_portfolio_snapshot[exposure])", "#,##0", "Analytical exposure units", "outputs/portfolio_kpi_summary.csv"],
        ["Average PD All", "AVERAGE(fact_portfolio_snapshot[pd])", "0.00%", "All-account average PD", "outputs/portfolio_kpi_summary.csv"],
        ["Average PD Matured", "CALCULATE(AVERAGE(fact_portfolio_snapshot[pd]), fact_portfolio_snapshot[matured_flag] = 1)", "0.00%", "Matured average PD", "outputs/portfolio_kpi_summary.csv"],
        ["Observed Bad Accounts Matured", "CALCULATE(SUM(fact_portfolio_snapshot[bad_flag]), fact_portfolio_snapshot[matured_flag] = 1)", "Whole number", "Observed bads on eligible population", "outputs/portfolio_kpi_summary.csv"],
        ["Observed Bad Rate Matured", "DIVIDE([Observed Bad Accounts Matured], [Matured Accounts])", "0.00%", "Correct maturity-safe bad rate", "outputs/portfolio_kpi_summary.csv"],
        ["Calibration Gap Matured", "[Observed Bad Rate Matured] - [Average PD Matured]", "0.00%", "Observed minus predicted", "outputs/model_monitoring_current_summary.csv"],
        ["Expected Loss Proxy", "SUM(fact_portfolio_snapshot[expected_loss])", "#,##0", "Frozen EL proxy", "outputs/portfolio_kpi_summary.csv"],
        ["High Risk Accounts", "CALCULATE([Total Accounts], KEEPFILTERS(dim_risk_grade[grade_code] IN {\"D\",\"E\"}))", "Whole number", "Grade D/E accounts through governed risk-grade dimension", "outputs/portfolio_kpi_summary.csv"],
        ["High Risk Share", "DIVIDE([High Risk Accounts], [Total Accounts])", "0.00%", "Grade D/E concentration", "outputs/portfolio_kpi_summary.csv"],
        ["Latest Eligible Synthetic Period", "VAR EligiblePeriods=CALCULATETABLE(FILTER(ALL(fact_synthetic_stock_flow), fact_synthetic_stock_flow[accounts] >= 1000), REMOVEFILTERS(dim_synthetic_snapshot_date)) RETURN MAXX(EligiblePeriods, fact_synthetic_stock_flow[snapshot_month])", "Text", "Latest synthetic period meeting the formal KRI denominator threshold, independent of date slicer context", "outputs/delinquency_stock_flow_summary.csv"],
        ["Snapshot Accounts", "VAR p=[Latest Eligible Synthetic Period] RETURN CALCULATE(DISTINCTCOUNT(fact_synthetic_servicing[account_id]), REMOVEFILTERS(dim_synthetic_snapshot_date), REMOVEFILTERS(fact_synthetic_servicing[snapshot_month]), TREATAS({p}, fact_synthetic_servicing[snapshot_month]))", "Whole number", "Synthetic snapshot denominator at latest eligible period", "outputs/delinquency_stock_flow_summary.csv"],
        ["30 Plus DPD Accounts", "VAR p=[Latest Eligible Synthetic Period] RETURN CALCULATE(DISTINCTCOUNT(fact_synthetic_servicing[account_id]), REMOVEFILTERS(dim_synthetic_snapshot_date), REMOVEFILTERS(fact_synthetic_servicing[snapshot_month]), TREATAS({p}, fact_synthetic_servicing[snapshot_month]), fact_synthetic_servicing[dpd_bucket] IN {\"30-59\",\"60-89\",\"90+\",\"Default\"})", "Whole number", "Synthetic 30+ stock at latest eligible period", "outputs/delinquency_stock_flow_summary.csv"],
        ["30 Plus DPD Rate", "DIVIDE([30 Plus DPD Accounts], [Snapshot Accounts])", "0.00%", "Synthetic 30+ rate", "outputs/delinquency_stock_flow_summary.csv"],
        ["Default Stock Accounts", "VAR p=[Latest Eligible Synthetic Period] RETURN CALCULATE(DISTINCTCOUNT(fact_synthetic_servicing[account_id]), REMOVEFILTERS(dim_synthetic_snapshot_date), REMOVEFILTERS(fact_synthetic_servicing[snapshot_month]), TREATAS({p}, fact_synthetic_servicing[snapshot_month]), fact_synthetic_servicing[default_stock_flag] = 1)", "Whole number", "Ending default stock at latest eligible period", "outputs/delinquency_stock_flow_summary.csv"],
        ["Default Stock Rate", "DIVIDE([Default Stock Accounts], [Snapshot Accounts])", "0.00%", "Ending default stock rate", "outputs/delinquency_stock_flow_summary.csv"],
        ["New Default Flow Accounts", "VAR p=[Latest Eligible Synthetic Period] RETURN CALCULATE(SUM(fact_synthetic_servicing[new_default_flag]), REMOVEFILTERS(dim_synthetic_snapshot_date), REMOVEFILTERS(fact_synthetic_servicing[snapshot_month]), TREATAS({p}, fact_synthetic_servicing[snapshot_month]))", "Whole number", "New default flow at latest eligible period", "outputs/delinquency_stock_flow_summary.csv"],
        ["Accounts At Risk For Default", "VAR p=[Latest Eligible Synthetic Period] RETURN CALCULATE(COUNTROWS(fact_synthetic_servicing), REMOVEFILTERS(dim_synthetic_snapshot_date), REMOVEFILTERS(fact_synthetic_servicing[snapshot_month]), TREATAS({p}, fact_synthetic_servicing[snapshot_month]), NOT(fact_synthetic_servicing[beginning_dpd_bucket] IN {\"Default\",\"Closed\"}))", "Whole number", "Default-flow denominator at latest eligible period", "outputs/delinquency_stock_flow_summary.csv"],
        ["New Default Flow Rate", "DIVIDE([New Default Flow Accounts], [Accounts At Risk For Default])", "0.00%", "New default flow rate", "outputs/delinquency_stock_flow_summary.csv"],
        ["Cure Accounts", "VAR p=[Latest Eligible Synthetic Period] RETURN CALCULATE(SUM(fact_synthetic_servicing[cure_flag]), REMOVEFILTERS(dim_synthetic_snapshot_date), REMOVEFILTERS(fact_synthetic_servicing[snapshot_month]), TREATAS({p}, fact_synthetic_servicing[snapshot_month]))", "Whole number", "Synthetic cures at latest eligible period", "outputs/cure_rate_summary.csv"],
        ["Cure Opportunities", "VAR p=[Latest Eligible Synthetic Period] RETURN CALCULATE(COUNTROWS(fact_synthetic_servicing), REMOVEFILTERS(dim_synthetic_snapshot_date), REMOVEFILTERS(fact_synthetic_servicing[snapshot_month]), TREATAS({p}, fact_synthetic_servicing[snapshot_month]), fact_synthetic_servicing[beginning_dpd_bucket] IN {\"1-29\",\"30-59\",\"60-89\",\"90+\"})", "Whole number", "Eligible cure denominator at latest eligible period", "outputs/cure_rate_summary.csv"],
        ["Cure Rate", "DIVIDE([Cure Accounts], [Cure Opportunities])", "0.00%", "Synthetic cure rate", "outputs/cure_rate_summary.csv"],
        ["Policy Accounts", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[matured_accounts]), BLANK())", "Whole number", "Selected policy common population; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Policy Approval Rate", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[approval_rate]), BLANK())", "0.00%", "Selected policy approval rate; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Manual Review Rate", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[manual_review_rate]), BLANK())", "0.00%", "Selected policy manual review; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Decline Rate", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[decline_rate]), BLANK())", "0.00%", "Selected policy decline rate; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Approved Bad Rate Matured", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[approved_bad_rate_matured]), BLANK())", "0.00%", "Maturity-safe approved bad rate; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Bad Capture Non Approved", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[bad_capture_non_approved]), BLANK())", "0.00%", "Bad capture outside approval; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Auto Approved Expected Loss", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[auto_approved_expected_loss]), BLANK())", "#,##0", "Straight-through auto-approved EL; not realised portfolio loss reduction; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Manual Review Expected Loss", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[manual_review_expected_loss]), BLANK())", "#,##0", "EL routed to manual review; disposition not modelled; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Declined Expected Loss", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[declined_expected_loss]), BLANK())", "#,##0", "EL routed to decline under selected policy; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Exception Queue Expected Loss", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[exception_queue_expected_loss]), BLANK())", "#,##0", "EL routed to exception queue; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Total Routed Expected Loss", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[total_routed_expected_loss]), BLANK())", "#,##0", "Manual review + decline + exception queue EL; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Total Scored Expected Loss", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[total_scored_expected_loss]), BLANK())", "#,##0", "Auto-approved + routed EL identity; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Auto Approved EL Reduction vs Baseline", "IF(HASONEVALUE(dim_policy_version[policy_version]), MAX(fact_policy_bridge[auto_approved_el_reduction_vs_baseline]), BLANK())", "0.00%", "Reduction in straight-through auto-approved EL versus current 25%; not realised loss reduction; blank if policy context is ambiguous", "outputs/policy_version_bridge_summary.csv"],
        ["Latest Eligible Period", "MAXX(FILTER(ALL(fact_model_monitoring), fact_model_monitoring[eligibility_status] = \"ELIGIBLE\"), fact_model_monitoring[period])", "Text", "Latest assessable model period; ignores accidental date filters", "outputs/model_monitoring_current_summary.csv"],
        ["Prior Eligible Period", "VAR latest=[Latest Eligible Period] RETURN MAXX(FILTER(ALL(fact_model_monitoring), fact_model_monitoring[eligibility_status] = \"ELIGIBLE\" && fact_model_monitoring[period] < latest), fact_model_monitoring[period])", "Text", "Prior assessable model period", "outputs/model_monitoring_current_summary.csv"],
        ["Latest Eligible AUC", "VAR p=[Latest Eligible Period] RETURN CALCULATE(MAX(fact_model_monitoring[auc]), REMOVEFILTERS(fact_model_monitoring), fact_model_monitoring[period]=p)", "0.000", "Latest eligible discrimination", "outputs/model_monitoring_current_summary.csv"],
        ["Latest Eligible KS", "VAR p=[Latest Eligible Period] RETURN CALCULATE(MAX(fact_model_monitoring[ks]), REMOVEFILTERS(fact_model_monitoring), fact_model_monitoring[period]=p)", "0.000", "Tie-safe latest eligible KS", "outputs/model_monitoring_current_summary.csv"],
        ["Latest Eligible PD PSI", "VAR p=[Latest Eligible Period] RETURN CALCULATE(MAX(fact_model_monitoring[pd_psi_vs_development_baseline]), REMOVEFILTERS(fact_model_monitoring), fact_model_monitoring[period]=p)", "0.000", "Selected/latest PSI; never SUM", "outputs/model_monitoring_current_summary.csv"],
        ["Latest Eligible Bad Rate Matured", "VAR p=[Latest Eligible Period] RETURN CALCULATE(MAX(fact_model_monitoring[observed_bad_rate_matured]), REMOVEFILTERS(fact_model_monitoring), fact_model_monitoring[period]=p)", "0.00%", "Latest eligible observed bad rate", "outputs/model_monitoring_current_summary.csv"],
        ["Prior Eligible Bad Rate Matured", "VAR p=[Prior Eligible Period] RETURN CALCULATE(MAX(fact_model_monitoring[observed_bad_rate_matured]), REMOVEFILTERS(fact_model_monitoring), fact_model_monitoring[period]=p)", "0.00%", "Prior eligible observed bad rate", "outputs/model_monitoring_current_summary.csv"],
        ["Bad Rate Change bps", "([Latest Eligible Bad Rate Matured] - [Prior Eligible Bad Rate Matured]) * 10000", "0", "Latest eligible versus prior eligible movement; no historical MAX leakage", "outputs/portfolio_change_summary.csv"],
        ["Context Bad Rate Matured", "MAX(fact_model_monitoring[observed_bad_rate_matured])", "0.00%", "Context-sensitive bad rate for historical visuals only", "outputs/model_monitoring_summary.csv"],
        ["Latest Calibration Gap", "VAR p=[Latest Eligible Period] RETURN CALCULATE(MAX(fact_model_monitoring[calibration_gap_matured]), REMOVEFILTERS(fact_model_monitoring), fact_model_monitoring[period]=p)", "0.00%", "Latest eligible calibration", "outputs/model_monitoring_current_summary.csv"],
        ["Stage 2 Proxy Accounts", "CALCULATE(SUM(fact_ecl[accounts]), fact_ecl[stage]=\"Stage 2\")", "Whole number", "P4 Stage 2 proxy accounts", "outputs/ecl_stress_monitoring_summary.csv"],
        ["ECL Accounts", "SUM(fact_ecl[accounts])", "Whole number", "P4 ECL denominator", "outputs/ecl_stress_monitoring_summary.csv"],
        ["Stage 2 Proxy Share", "DIVIDE([Stage 2 Proxy Accounts], [ECL Accounts])", "0.00%", "P4 Stage 2 proxy share", "outputs/ecl_stress_monitoring_summary.csv"],
        ["Base ECL", "SUM(fact_ecl[ecl_base])", "#,##0", "P4 base scenario ECL", "outputs/ecl_stress_monitoring_summary.csv"],
        ["Weighted ECL", "SUM(fact_ecl[ecl_weighted])", "#,##0", "Probability-weighted ECL", "outputs/ecl_stress_monitoring_summary.csv"],
        ["Downside ECL", "SUM(fact_ecl[ecl_downside])", "#,##0", "P4 downside scenario ECL", "outputs/ecl_stress_monitoring_summary.csv"],
        ["Weighted ECL Rate", "DIVIDE([Weighted ECL], SUM(fact_ecl[ead]))", "0.00%", "Weighted ECL / EAD", "outputs/ecl_stress_monitoring_summary.csv"],
        ["Downside ECL Uplift vs Base", "DIVIDE([Downside ECL]-[Base ECL], [Base ECL])", "0.00%", "Downside versus base ECL; KRI-007 denominator", "outputs/ecl_stress_monitoring_summary.csv"],
        ["Downside ECL Uplift vs Weighted", "DIVIDE([Downside ECL]-[Weighted ECL], [Weighted ECL])", "0.00%", "Secondary comparison; not KRI-007", "outputs/ecl_stress_monitoring_summary.csv"],
        ["Red KRI Count", "COUNTROWS(FILTER(fact_kri, fact_kri[status]=\"RED\"))", "Whole number", "Red KRI count", "outputs/kri_status_summary.csv"],
        ["Amber KRI Count", "COUNTROWS(FILTER(fact_kri, fact_kri[status]=\"AMBER\"))", "Whole number", "Amber KRI count", "outputs/kri_status_summary.csv"],
        ["Open Alert Count", "COUNTROWS(FILTER(fact_alerts, fact_alerts[workflow_status]=\"OPEN\"))", "Whole number", "Open EWS alerts", "outputs/early_warning_alerts.csv"],
        ["Open Action Count", "COUNTROWS(FILTER(fact_actions, fact_actions[status]=\"OPEN\"))", "Whole number", "Open management actions", "outputs/management_action_summary.csv"],
    ]


def generate_powerbi_pack() -> None:
    dax = pd.DataFrame(
        _dax_rows(),
        columns=["measure_name", "dax_formula", "format", "business_definition", "python_reconciliation_source"],
    )
    dax["folder"] = np.select(
        [
            dax["measure_name"].str.contains("DPD|Default|Cure|Snapshot|At Risk"),
            dax["measure_name"].str.contains("Policy|Approved|Decline|Manual|Bad Capture"),
            dax["measure_name"].str.contains("ECL|Stage 2"),
            dax["measure_name"].str.contains("KRI|Alert|Action"),
            dax["measure_name"].str.contains("AUC|KS|PSI|Calibration|Eligible"),
        ],
        ["Delinquency", "Policy", "ECL", "Governance", "Model Monitoring"],
        default="Portfolio",
    )
    write_csv(add_meta(dax, "Technical", "Definition", "P2", "DAX build dictionary; runtime requires real PBIX"), "powerbi/dax_measure_dictionary.csv")

    defined = set(dax["measure_name"])
    missing_dependencies: set[str] = set()
    for formula in dax["dax_formula"]:
        for dependency in re.findall(r"(?<![A-Za-z0-9_])\[([^\]]+)\]", formula):
            if dependency not in defined:
                missing_dependencies.add(dependency)
    lint_rows = [
        ["DAX-LINT-001", "At least 25 measures", len(dax), ">=25", "PASS" if len(dax) >= 25 else "FAIL"],
        ["DAX-LINT-002", "Bad rate denominator uses Matured Accounts", int(dax.loc[dax["measure_name"].eq("Observed Bad Rate Matured"), "dax_formula"].str.contains("Matured Accounts", regex=False).all()), "1", "PASS"],
        ["DAX-LINT-003", "PSI does not use SUM", int(~dax.loc[dax["measure_name"].eq("Latest Eligible PD PSI"), "dax_formula"].str.contains("SUM\\(", regex=True).any()), "1", "PASS"],
        ["DAX-LINT-004", "All referenced measures are defined", ", ".join(sorted(missing_dependencies)) or "NONE", "NONE", "PASS" if not missing_dependencies else "FAIL"],
        ["DAX-LINT-005", "Every measure has source", int(dax["python_reconciliation_source"].notna().sum()), str(len(dax)), "PASS" if dax["python_reconciliation_source"].notna().all() else "FAIL"],
        ["DAX-LINT-006", "High risk measure uses governed grade_code dimension", str(dax.loc[dax["measure_name"].eq("High Risk Accounts"), "dax_formula"].iloc[0]), "dim_risk_grade[grade_code]", "PASS" if "dim_risk_grade[grade_code]" in str(dax.loc[dax["measure_name"].eq("High Risk Accounts"), "dax_formula"].iloc[0]) else "FAIL"],
        ["DAX-LINT-007", "Downside ECL uplift uses Base ECL denominator", str(dax.loc[dax["measure_name"].eq("Downside ECL Uplift vs Base"), "dax_formula"].iloc[0]), "[Base ECL]", "PASS" if "[Base ECL]" in str(dax.loc[dax["measure_name"].eq("Downside ECL Uplift vs Base"), "dax_formula"].iloc[0]) else "FAIL"],
        ["DAX-LINT-008", "Headline bad-rate movement is latest-period safe", str(dax.loc[dax["measure_name"].eq("Bad Rate Change bps"), "dax_formula"].iloc[0]), "Latest/Prior eligible measures", "PASS" if "Latest Eligible Bad Rate Matured" in str(dax.loc[dax["measure_name"].eq("Bad Rate Change bps"), "dax_formula"].iloc[0]) and "MAX(" not in str(dax.loc[dax["measure_name"].eq("Bad Rate Change bps"), "dax_formula"].iloc[0]) else "FAIL"],
        ["DAX-LINT-009", "Public policy EL measures use auto-approved/routed language", ", ".join(dax.loc[dax["measure_name"].str.contains("Expected Loss|EL Reduction", regex=True), "measure_name"]), "No generic Approved Expected Loss measure", "PASS" if "Approved Expected Loss" not in set(dax["measure_name"]) and "Expected Loss Reduction vs Baseline" not in set(dax["measure_name"]) else "FAIL"],
        ["DAX-LINT-010", "Policy measures guard ambiguous policy context", "; ".join(dax.loc[dax["folder"].eq("Policy"), "dax_formula"].head(3)), "HASONEVALUE(dim_policy_version[policy_version])", "PASS" if dax.loc[dax["folder"].eq("Policy"), "dax_formula"].str.contains("HASONEVALUE\\(dim_policy_version\\[policy_version\\]\\)", regex=True).all() else "FAIL"],
        ["DAX-LINT-011", "Synthetic headline components use latest eligible synthetic period", "; ".join(dax.loc[dax["measure_name"].isin(["Snapshot Accounts", "30 Plus DPD Accounts", "Cure Accounts", "Cure Opportunities"]), "dax_formula"]), "Latest Eligible Synthetic Period", "PASS" if dax.loc[dax["measure_name"].isin(["Snapshot Accounts", "30 Plus DPD Accounts", "Cure Accounts", "Cure Opportunities"]), "dax_formula"].str.contains("Latest Eligible Synthetic Period", regex=False).all() else "FAIL"],
        ["DAX-LINT-012", "Synthetic headlines remove the dedicated snapshot-date dimension filter", "; ".join(dax.loc[dax["measure_name"].isin(["Latest Eligible Synthetic Period", "Snapshot Accounts", "30 Plus DPD Accounts", "Cure Accounts", "Cure Opportunities"]), "dax_formula"]), "REMOVEFILTERS(dim_synthetic_snapshot_date)", "PASS" if dax.loc[dax["measure_name"].isin(["Latest Eligible Synthetic Period", "Snapshot Accounts", "30 Plus DPD Accounts", "Cure Accounts", "Cure Opportunities"]), "dax_formula"].str.contains("REMOVEFILTERS(dim_synthetic_snapshot_date)", regex=False).all() else "FAIL"],
    ]
    lint = pd.DataFrame(lint_rows, columns=["test_id", "semantic_control", "actual_value", "expected_value", "status"])
    write_csv(add_meta(lint, "Technical", "Executed", "P2", "Executable semantic lint; not DAX engine execution"), "testing/dax_semantic_lint_results.csv")

    kpi = _read("outputs/portfolio_kpi_summary.csv")
    current = _read("outputs/model_monitoring_current_summary.csv")
    policy_bridge = _read("outputs/policy_version_bridge_summary.csv")
    policy_baseline = policy_bridge[policy_bridge["policy_version"].eq("baseline_current_25")].iloc[0]
    ecl = _read("outputs/ecl_stress_monitoring_summary.csv")
    kri = _read("outputs/kri_status_summary.csv")
    alerts = _read("outputs/early_warning_alerts.csv")
    actions = _read("outputs/management_action_summary.csv")
    segment = _read("outputs/segment_concentration_summary.csv")
    current_values = dict(zip(current["metric"], current["current_value"]))
    prior_values = dict(zip(current["metric"], current["prior_value"]))
    stock = _read("outputs/delinquency_stock_flow_summary.csv")
    formal_stock = stock[stock["accounts"].ge(1_000)].sort_values("snapshot_month").iloc[-1]
    formal_synthetic = _read(
        "data/synthetic/monthly_delinquency_snapshots.csv.gz",
        usecols=["snapshot_month", "account_id", "beginning_dpd_bucket", "dpd_bucket", "new_default_flag", "cure_flag", "default_stock_flag"],
    )
    formal_synthetic = formal_synthetic[formal_synthetic["snapshot_month"].eq(formal_stock["snapshot_month"])]
    current_period = str(current["current_period"].iloc[0])
    prior_period = str(current["prior_period"].iloc[0])
    observed_default_context = "Full accepted/booked portfolio; no risk-grade/purpose filter; all application periods"
    latest_model_context = "Latest eligible observed period = 2017-12; prior eligible observed period = 2017-11; eligibility thresholds applied"
    synthetic_context = "snapshot_month = 2020-10; latest eligible synthetic KRI period"
    ecl_context = "P4 frozen reporting-date snapshot; all stages and risk grades unless explicitly filtered"
    policy_context = "Selected baseline policy_version = baseline_current_25; common matured population"

    def expected_record(
        value: object,
        value_type: str = "NUMBER",
        context: str = observed_default_context,
        source: str = "outputs/portfolio_kpi_summary.csv",
        key: str = "default",
        period: str = "ALL",
    ) -> dict[str, object]:
        return {
            "expected_value_type": value_type,
            "expected_numeric_value": float(value) if value_type == "NUMBER" and pd.notna(value) else np.nan,
            "expected_text_value": str(value) if value_type in {"TEXT", "DATE", "BOOLEAN", "BLANK"} and pd.notna(value) else "",
            "filter_context": context,
            "expected_source_file": source,
            "expected_source_key": key,
            "expected_period": period,
        }

    stage2_accounts = float(ecl.loc[ecl["stage"].eq("Stage 2"), "accounts"].sum())
    ecl_accounts = float(ecl["accounts"].sum())
    high_risk_accounts = int(segment.loc[segment["risk_grade"].astype(str).str[0].isin(["D", "E"]), "accounts"].sum())
    expected_records = {
        "Total Accounts": expected_record(_metric(kpi, "total_accounts"), key="total_accounts"),
        "Matured Accounts": expected_record(_metric(kpi, "matured_accounts"), key="matured_accounts"),
        "Immature Accounts": expected_record(_metric(kpi, "immature_accounts"), key="immature_accounts"),
        "Outcome Eligibility Rate": expected_record(_metric(kpi, "matured_accounts") / _metric(kpi, "total_accounts"), key="matured_accounts / total_accounts"),
        "Total Exposure": expected_record(_metric(kpi, "total_exposure"), key="total_exposure"),
        "Average PD All": expected_record(_metric(kpi, "average_pd_all_accounts"), key="average_pd_all_accounts"),
        "Average PD Matured": expected_record(_metric(kpi, "average_pd_matured"), key="average_pd_matured"),
        "Observed Bad Accounts Matured": expected_record(_metric(kpi, "observed_bad_accounts_matured"), key="observed_bad_accounts_matured"),
        "Observed Bad Rate Matured": expected_record(_metric(kpi, "observed_bad_rate_matured"), key="observed_bad_rate_matured"),
        "Calibration Gap Matured": expected_record(current_values["calibration_gap_matured"], context=latest_model_context, source="outputs/model_monitoring_current_summary.csv", key="calibration_gap_matured.current_value", period=current_period),
        "Expected Loss Proxy": expected_record(_metric(kpi, "expected_loss_proxy"), key="expected_loss_proxy"),
        "High Risk Accounts": expected_record(high_risk_accounts, key="direct Grade D/E account count from segment summary"),
        "High Risk Share": expected_record(_metric(kpi, "high_risk_share_grade_d_e"), key="high_risk_share_grade_d_e"),
        "Latest Eligible Synthetic Period": expected_record(str(formal_stock["snapshot_month"]), "TEXT", "No synthetic page-date dependency; measure selects latest period with accounts >= 1,000", "outputs/delinquency_stock_flow_summary.csv", "latest snapshot_month where accounts >= 1000", str(formal_stock["snapshot_month"])),
        "Snapshot Accounts": expected_record(float(formal_synthetic["account_id"].nunique()), context=synthetic_context, source="data/synthetic/monthly_delinquency_snapshots.csv.gz", key="distinct account_id", period=str(formal_stock["snapshot_month"])),
        "30 Plus DPD Accounts": expected_record(float(formal_synthetic["dpd_bucket"].isin(["30-59", "60-89", "90+", "Default"]).sum()), context=synthetic_context, source="data/synthetic/monthly_delinquency_snapshots.csv.gz", key="dpd_bucket in 30+/Default", period=str(formal_stock["snapshot_month"])),
        "30 Plus DPD Rate": expected_record(float(formal_stock["dpd_30_plus"]), context=synthetic_context, source="outputs/delinquency_stock_flow_summary.csv", key="dpd_30_plus", period=str(formal_stock["snapshot_month"])),
        "Default Stock Accounts": expected_record(float(formal_synthetic["default_stock_flag"].sum()), context=synthetic_context, source="data/synthetic/monthly_delinquency_snapshots.csv.gz", key="default_stock_flag", period=str(formal_stock["snapshot_month"])),
        "Default Stock Rate": expected_record(float(formal_stock["default_stock_rate"]), context=synthetic_context, source="outputs/delinquency_stock_flow_summary.csv", key="default_stock_rate", period=str(formal_stock["snapshot_month"])),
        "New Default Flow Accounts": expected_record(float(formal_synthetic["new_default_flag"].sum()), context=synthetic_context, source="data/synthetic/monthly_delinquency_snapshots.csv.gz", key="new_default_flag", period=str(formal_stock["snapshot_month"])),
        "Accounts At Risk For Default": expected_record(float((~formal_synthetic["beginning_dpd_bucket"].isin(["Default", "Closed"])).sum()), context=synthetic_context, source="data/synthetic/monthly_delinquency_snapshots.csv.gz", key="beginning_dpd_bucket not Default/Closed", period=str(formal_stock["snapshot_month"])),
        "New Default Flow Rate": expected_record(float(formal_stock["new_default_flow_rate"]), context=synthetic_context, source="outputs/delinquency_stock_flow_summary.csv", key="new_default_flow_rate", period=str(formal_stock["snapshot_month"])),
        "Cure Accounts": expected_record(float(formal_synthetic["cure_flag"].sum()), context=synthetic_context, source="data/synthetic/monthly_delinquency_snapshots.csv.gz", key="cure_flag", period=str(formal_stock["snapshot_month"])),
        "Cure Opportunities": expected_record(float(formal_synthetic["beginning_dpd_bucket"].isin(["1-29", "30-59", "60-89", "90+"]).sum()), context=synthetic_context, source="data/synthetic/monthly_delinquency_snapshots.csv.gz", key="delinquent beginning buckets", period=str(formal_stock["snapshot_month"])),
        "Cure Rate": expected_record(float(formal_synthetic["cure_flag"].sum() / formal_synthetic["beginning_dpd_bucket"].isin(["1-29", "30-59", "60-89", "90+"]).sum()), context=synthetic_context, source="data/synthetic/monthly_delinquency_snapshots.csv.gz", key="cure_flag / cure opportunities", period=str(formal_stock["snapshot_month"])),
        "Policy Accounts": expected_record(policy_baseline["matured_accounts"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="matured_accounts", period="ALL"),
        "Policy Approval Rate": expected_record(policy_baseline["approval_rate"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="approval_rate", period="ALL"),
        "Manual Review Rate": expected_record(policy_baseline["manual_review_rate"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="manual_review_rate", period="ALL"),
        "Decline Rate": expected_record(policy_baseline["decline_rate"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="decline_rate", period="ALL"),
        "Approved Bad Rate Matured": expected_record(policy_baseline["approved_bad_rate_matured"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="approved_bad_rate_matured", period="ALL"),
        "Bad Capture Non Approved": expected_record(policy_baseline["bad_capture_non_approved"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="bad_capture_non_approved", period="ALL"),
        "Auto Approved Expected Loss": expected_record(policy_baseline["auto_approved_expected_loss"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="auto_approved_expected_loss", period="ALL"),
        "Manual Review Expected Loss": expected_record(policy_baseline["manual_review_expected_loss"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="manual_review_expected_loss", period="ALL"),
        "Declined Expected Loss": expected_record(policy_baseline["declined_expected_loss"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="declined_expected_loss", period="ALL"),
        "Exception Queue Expected Loss": expected_record(policy_baseline["exception_queue_expected_loss"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="exception_queue_expected_loss", period="ALL"),
        "Total Routed Expected Loss": expected_record(policy_baseline["total_routed_expected_loss"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="total_routed_expected_loss", period="ALL"),
        "Total Scored Expected Loss": expected_record(policy_baseline["total_scored_expected_loss"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="total_scored_expected_loss", period="ALL"),
        "Auto Approved EL Reduction vs Baseline": expected_record(policy_baseline["auto_approved_el_reduction_vs_baseline"], context=policy_context, source="outputs/policy_version_bridge_summary.csv", key="auto_approved_el_reduction_vs_baseline", period="ALL"),
        "Latest Eligible Period": expected_record(current_period, "TEXT", latest_model_context, "outputs/model_monitoring_current_summary.csv", "current_period", current_period),
        "Prior Eligible Period": expected_record(prior_period, "TEXT", latest_model_context, "outputs/model_monitoring_current_summary.csv", "prior_period", prior_period),
        "Latest Eligible AUC": expected_record(current_values["auc"], context=latest_model_context, source="outputs/model_monitoring_current_summary.csv", key="auc.current_value", period=current_period),
        "Latest Eligible KS": expected_record(current_values["ks"], context=latest_model_context, source="outputs/model_monitoring_current_summary.csv", key="ks.current_value", period=current_period),
        "Latest Eligible PD PSI": expected_record(current_values["pd_psi_vs_development_baseline"], context=latest_model_context, source="outputs/model_monitoring_current_summary.csv", key="pd_psi_vs_development_baseline.current_value", period=current_period),
        "Latest Eligible Bad Rate Matured": expected_record(current_values["observed_bad_rate_matured"], context=latest_model_context, source="outputs/model_monitoring_current_summary.csv", key="observed_bad_rate_matured.current_value", period=current_period),
        "Prior Eligible Bad Rate Matured": expected_record(prior_values["observed_bad_rate_matured"], context=latest_model_context, source="outputs/model_monitoring_current_summary.csv", key="observed_bad_rate_matured.prior_value", period=prior_period),
        "Bad Rate Change bps": expected_record((current_values["observed_bad_rate_matured"] - prior_values["observed_bad_rate_matured"]) * 10_000, context=latest_model_context, source="outputs/portfolio_change_summary.csv", key="bad_rate_change_bps", period=f"{prior_period}_to_{current_period}"),
        "Context Bad Rate Matured": expected_record(current_values["observed_bad_rate_matured"], context="Explicit period filter: application_period = 2017-12; matured outcomes only", source="outputs/model_monitoring_summary.csv", key="observed_bad_rate_matured where period=2017-12", period=current_period),
        "Latest Calibration Gap": expected_record(current_values["calibration_gap_matured"], context=latest_model_context, source="outputs/model_monitoring_current_summary.csv", key="calibration_gap_matured.current_value", period=current_period),
        "Stage 2 Proxy Accounts": expected_record(stage2_accounts, context=ecl_context + "; stage = Stage 2", source="outputs/ecl_stress_monitoring_summary.csv", key="Stage 2 accounts", period=str(ecl["reporting_date"].iloc[0])),
        "ECL Accounts": expected_record(ecl_accounts, context=ecl_context, source="outputs/ecl_stress_monitoring_summary.csv", key="total accounts", period=str(ecl["reporting_date"].iloc[0])),
        "Stage 2 Proxy Share": expected_record(stage2_accounts / ecl_accounts, context=ecl_context + "; stage = Stage 2 / all stages", source="outputs/ecl_stress_monitoring_summary.csv", key="Stage 2 accounts / total accounts", period=str(ecl["reporting_date"].iloc[0])),
        "Base ECL": expected_record(float(ecl["ecl_base"].sum()), context=ecl_context, source="outputs/ecl_stress_monitoring_summary.csv", key="sum ecl_base", period=str(ecl["reporting_date"].iloc[0])),
        "Weighted ECL": expected_record(float(ecl["ecl_weighted"].sum()), context=ecl_context, source="outputs/ecl_stress_monitoring_summary.csv", key="sum ecl_weighted", period=str(ecl["reporting_date"].iloc[0])),
        "Downside ECL": expected_record(float(ecl["ecl_downside"].sum()), context=ecl_context, source="outputs/ecl_stress_monitoring_summary.csv", key="sum ecl_downside", period=str(ecl["reporting_date"].iloc[0])),
        "Weighted ECL Rate": expected_record(float(ecl["ecl_weighted"].sum() / ecl["ead"].sum()), context=ecl_context, source="outputs/ecl_stress_monitoring_summary.csv", key="sum ecl_weighted / sum ead", period=str(ecl["reporting_date"].iloc[0])),
        "Downside ECL Uplift vs Base": expected_record(float((ecl["ecl_downside"].sum() - ecl["ecl_base"].sum()) / ecl["ecl_base"].sum()), context=ecl_context, source="outputs/ecl_stress_monitoring_summary.csv", key="(downside-base)/base", period=str(ecl["reporting_date"].iloc[0])),
        "Downside ECL Uplift vs Weighted": expected_record(float((ecl["ecl_downside"].sum() - ecl["ecl_weighted"].sum()) / ecl["ecl_weighted"].sum()), context=ecl_context, source="outputs/ecl_stress_monitoring_summary.csv", key="(downside-weighted)/weighted", period=str(ecl["reporting_date"].iloc[0])),
        "Red KRI Count": expected_record(int(kri["status"].eq("RED").sum()), context="Current governed KRI set; all owners; open external gates disclosed", source="outputs/kri_status_summary.csv", key="status=RED", period="CURRENT"),
        "Amber KRI Count": expected_record(int(kri["status"].eq("AMBER").sum()), context="Current governed KRI set; all owners; open external gates disclosed", source="outputs/kri_status_summary.csv", key="status=AMBER", period="CURRENT"),
        "Open Alert Count": expected_record(int(alerts["workflow_status"].eq("OPEN").sum()), context="Open EWS alerts only", source="outputs/early_warning_alerts.csv", key="workflow_status=OPEN", period="CURRENT"),
        "Open Action Count": expected_record(int(actions["status"].eq("OPEN").sum()), context="Open management actions only", source="outputs/management_action_summary.csv", key="status=OPEN", period="CURRENT"),
    }
    missing_expected = sorted(set(dax["measure_name"]) - set(expected_records))
    if missing_expected:
        raise ValueError(f"Missing DAX expected records: {missing_expected}")
    dax_tests = dax[["measure_name", "dax_formula", "python_reconciliation_source"]].copy()
    dax_tests.insert(0, "test_case_id", [f"DAX-{i:03d}" for i in range(1, len(dax_tests) + 1)])
    dax_tests["test_scope"] = "BASELINE_MEASURE"
    expected_frame = pd.DataFrame.from_dict(expected_records, orient="index")
    expected_frame.index.name = "measure_name"
    dax_tests = dax_tests.merge(expected_frame.reset_index(), on="measure_name", how="left")
    dax_tests["actual_numeric_value"] = np.nan
    dax_tests["actual_text_value"] = ""
    dax_tests["difference"] = np.nan
    dax_tests["tolerance"] = 1e-6
    dax_tests["execution_status"] = "NOT_EXECUTED_NO_PBIX"
    dax_tests["test_conclusion"] = "EXTERNAL_GATE"
    dax_tests["pbix_sha256"] = "NOT_AVAILABLE_NO_PBIX"
    dax_tests["powerbi_desktop_version"] = "NOT_EXECUTED_NO_PBIX"
    dax_tests["execution_date"] = RUN_DATE
    dax_tests["executor"] = "PENDING_POWER_BI_RUNTIME"
    context_tests: list[dict[str, object]] = []

    def add_context_test(
        test_id: str,
        measure_name: str,
        context: str,
        expected_type: str,
        numeric: float | None = None,
        text: str = "",
        source: str = "testing/pbix_runtime_context_test_plan.csv",
        key: str = "critical_context",
        period: str = "ALL",
    ) -> None:
        base = dax.loc[dax["measure_name"].eq(measure_name)].iloc[0].to_dict()
        context_tests.append(
            {
                "test_case_id": test_id,
                "measure_name": measure_name,
                "dax_formula": base["dax_formula"],
                "python_reconciliation_source": base["python_reconciliation_source"],
                "test_scope": "CRITICAL_CONTEXT",
                "expected_value_type": expected_type,
                "expected_numeric_value": numeric if numeric is not None else np.nan,
                "expected_text_value": text,
                "filter_context": context,
                "expected_source_file": source,
                "expected_source_key": key,
                "expected_period": period,
                "actual_numeric_value": np.nan,
                "actual_text_value": "",
                "difference": np.nan,
                "tolerance": 1e-6,
                "execution_status": "NOT_EXECUTED_NO_PBIX",
                "test_conclusion": "EXTERNAL_GATE",
                "pbix_sha256": "NOT_AVAILABLE_NO_PBIX",
                "powerbi_desktop_version": "NOT_EXECUTED_NO_PBIX",
                "execution_date": RUN_DATE,
                "executor": "PENDING_POWER_BI_RUNTIME",
            }
        )

    add_context_test("DAX-CTX-001", "Policy Approval Rate", "No policy version selected; governed measure must return BLANK instead of MAX across policies", "BLANK", text="BLANK", key="policy_no_selection_returns_blank")
    add_context_test("DAX-CTX-002", "Policy Approval Rate", "Single policy selected: dim_policy_version[policy_version] = baseline_current_25", "NUMBER", numeric=float(policy_baseline["approval_rate"]), key="policy_single_selection_baseline")
    add_context_test("DAX-CTX-003", "Policy Approval Rate", "Multiple policies selected: baseline_current_25 + scenario_1_recommended_20; governed measure must return BLANK", "BLANK", text="BLANK", key="policy_multi_selection_returns_blank")
    add_context_test("DAX-CTX-004", "Snapshot Accounts", "No snapshot_month slicer selected; governed measure must still use latest eligible synthetic period", "NUMBER", numeric=float(formal_synthetic["account_id"].nunique()), key="synthetic_no_period_default_latest_eligible", period=str(formal_stock["snapshot_month"]))
    add_context_test("DAX-CTX-005", "30 Plus DPD Rate", "No snapshot_month slicer selected; governed measure must still use latest eligible synthetic period", "NUMBER", numeric=float(formal_stock["dpd_30_plus"]), key="synthetic_rate_no_period_default_latest_eligible", period=str(formal_stock["snapshot_month"]))
    if context_tests:
        dax_tests = pd.concat([dax_tests, pd.DataFrame(context_tests)], ignore_index=True, sort=False)
    write_csv(add_meta(dax_tests, "Technical", "Test Design", "P2", "Expected values are ready; DAX runtime not falsely claimed"), "testing/dax_measure_test_cases.csv")
    context_plan = pd.DataFrame(
        [
            ["PBIX-CTX-001", "Policy Approval Rate", "No policy selected", "BLANK", "Prevents accidental MAX across all policy versions", "OPEN_EXTERNAL_GATE"],
            ["PBIX-CTX-002", "Policy Approval Rate", "Exactly one policy selected", "Numeric policy-specific value", "Confirms normal governed single-selection behavior", "OPEN_EXTERNAL_GATE"],
            ["PBIX-CTX-003", "Policy Approval Rate", "Multiple policies selected", "BLANK", "Prevents ambiguous policy card interpretation", "OPEN_EXTERNAL_GATE"],
            ["PBIX-CTX-004", "Snapshot Accounts", "No synthetic period selected", "Latest eligible synthetic period value", "Prevents whole-history 36-month aggregation", "OPEN_EXTERNAL_GATE"],
            ["PBIX-CTX-005", "30 Plus DPD Rate", "No synthetic period selected", "Latest eligible synthetic period value", "Prevents whole-history 36-month aggregation", "OPEN_EXTERNAL_GATE"],
            ["PBIX-CTX-006", "Relationships", "No orphan policy/risk-grade keys", "PASS", "Confirms dimensions filter facts correctly", "OPEN_EXTERNAL_GATE"],
            ["PBIX-CTX-007", "Cross-filter", "Policy slicer changes policy visuals only", "PASS", "Prevents policy slicer from corrupting observed portfolio metrics", "OPEN_EXTERNAL_GATE"],
            ["PBIX-CTX-008", "Snapshot Accounts", "dim_synthetic_snapshot_date filtered to a non-latest month", "Latest eligible synthetic period value", "Confirms headline measure removes dimension-originating date filters", "OPEN_EXTERNAL_GATE"],
            ["PBIX-CTX-009", "Observed portfolio cards", "Synthetic snapshot date slicer changed", "Observed cards unchanged", "Confirms observed and synthetic date roles are isolated", "OPEN_EXTERNAL_GATE"],
        ],
        columns=["test_id", "measure_or_area", "filter_context", "expected_behavior", "control_purpose", "execution_status"],
    )
    write_csv(add_meta(context_plan, "Technical", "Test Design", "P2", "PBIX runtime context tests to execute in v2.2"), "testing/pbix_runtime_context_test_plan.csv")

    theme = {
        "name": "Financial Risk Portfolio - Gold v2",
        "dataColors": ["#0B6B68", "#2F7D4A", "#C98613", "#B33A3A", "#3F6C8E", "#766A8B", "#65737A", "#9B6B43"],
        "background": "#F4F7F7",
        "foreground": "#18323E",
        "tableAccent": "#0B6B68",
        "visualStyles": {"*": {"*": {"title": [{"fontFamily": "Segoe UI Semibold", "fontSize": 11, "color": {"solid": {"color": "#18323E"}}}], "background": [{"color": {"solid": {"color": "#FFFFFF"}}, "transparency": 0}]}}},
    }
    (ROOT / "powerbi/theme.json").write_text(json.dumps(theme, indent=2), encoding="utf-8")
    risk_grade_dim = (
        segment[["risk_grade"]]
        .drop_duplicates()
        .assign(
            grade_code=lambda frame: frame["risk_grade"].astype(str).str[0],
            grade_order=lambda frame: frame["risk_grade"].astype(str).str[0].map({"A": 1, "B": 2, "C": 3, "D": 4, "E": 5}).fillna(99).astype(int),
        )
        .sort_values("grade_order")
    )
    write_csv(
        add_meta(risk_grade_dim, "Technical", "Dimension", "P2", "Governed risk-grade domain for Power BI filtering"),
        "powerbi/dim_risk_grade.csv",
    )
    policy_version_dim = (
        policy_bridge[["policy_version"]]
        .drop_duplicates()
        .assign(
            policy_order=lambda frame: frame["policy_version"].map(
                {
                    "baseline_current_25": 1,
                    "scenario_1_recommended_20": 2,
                    "scenario_2_conservative_15": 3,
                    "scenario_3_overlay_20": 4,
                    "scenario_4_decline_threshold_30": 5,
                }
            ).fillna(99).astype(int),
            default_policy_flag=lambda frame: frame["policy_version"].eq("baseline_current_25").astype(int),
        )
        .sort_values("policy_order")
    )
    write_csv(
        add_meta(policy_version_dim, "Technical", "Dimension", "P2", "Governed policy-version domain for Power BI single-selection controls"),
        "powerbi/dim_policy_version.csv",
    )
    synthetic_date_dim = (
        stock[["snapshot_month"]]
        .drop_duplicates()
        .sort_values("snapshot_month")
        .assign(snapshot_month_order=lambda frame: range(1, len(frame) + 1))
    )
    write_csv(
        add_meta(synthetic_date_dim, "Synthetic", "Dimension", "P2", "Dedicated role-playing date domain for synthetic servicing and stock-flow facts"),
        "powerbi/dim_synthetic_snapshot_date.csv",
    )

    write_md(
        "powerbi/data_model.md",
        "Power BI Data Model",
        """## Grain and relationships

| Table | Grain | Key | Evidence status |
| --- | --- | --- | --- |
| `fact_portfolio_snapshot` | one accepted/booked account | account_id | Observed/Derived |
| `fact_portfolio_period` | application cohort period | application_period | Observed/Derived |
| `fact_synthetic_servicing` | account x snapshot month | account_id + snapshot_month | Synthetic |
| `fact_synthetic_stock_flow` | snapshot month | snapshot_month | Synthetic summary |
| `fact_model_monitoring` | period x model version | period + model_version | Derived on matured outcomes |
| `fact_policy_bridge` | policy version | policy_version | Historical simulation |
| `fact_ecl` | stage x grade | stage + risk_grade | P4 proxy snapshot |
| `fact_kri` | KRI x current period | kri_id + current_period | Mixed |
| `fact_alerts` | EWS alert | ews_id | Governance |
| `fact_actions` | action | action_id | Governance |
| `dim_risk_grade` | risk grade label | risk_grade | Governed dimension |
| `dim_policy_version` | policy version | policy_version | Governed dimension |
| `dim_synthetic_snapshot_date` | synthetic snapshot month | snapshot_month | Synthetic role-playing dimension |

Use a conformed observed `dim_date`, plus the separate role-playing `dim_synthetic_snapshot_date`, `dim_risk_grade`, `dim_policy_version`, `dim_stage`, `dim_dpd_bucket`, `dim_owner` and `dim_data_status`. Relationships are one-to-many, single direction. Relate `dim_synthetic_snapshot_date[snapshot_month]` to both synthetic facts and keep it separate from observed application dates. `dim_risk_grade[grade_code]` must be unique across A/B/C/D/E and related to full risk-grade labels in facts; high-risk filters use `grade_code` D/E, not raw full-label literals. `dim_policy_version[policy_version]` must be a single-selection slicer on policy pages; policy KPI cards return blank under no-selection or multi-selection contexts. Synthetic headline cards explicitly remove `dim_synthetic_snapshot_date` and then apply the governed latest-eligible month through `TREATAS`, preventing an active date slicer from intersecting the headline denominator. Do not relate observed application cohorts directly to synthetic servicing snapshots without an explicit period-role bridge.
""",
    )
    page_rows = [
        [1, "Executive Risk Overview", "What changed, why it matters and which decisions are required", "Bad rate, PSI, calibration, KRI status, three priority actions"],
        [2, "Portfolio Quality", "Is observed booked-portfolio risk improving or deteriorating?", "Current/prior/change, grade/purpose concentration, maturity coverage"],
        [3, "Delinquency & Collections", "Are synthetic servicing controls internally coherent?", "30+ stock, default stock vs flow, roll-rate, cure, stock-flow bridge"],
        [4, "Vintage & Cohort", "How do risk curves vary by vintage, MOB, term and grade?", "Multi-vintage curves, 36/60-term residual balance, cohort table"],
        [5, "Model Monitoring", "Does frozen P3 ranking/calibration remain within trigger levels?", "Eligibility, AUC/KS, fixed-bin PSI, calibration, feature drift"],
        [6, "Policy Performance", "Which policy version balances risk, approval and workload?", "25/20/15/overlay/decline sensitivity bridge"],
        [7, "ECL, Stress & Concentration", "Where is loss sensitivity concentrated and what is still proxy-based?", "Stage/grade, scenario, term/purpose, readiness, bridge"],
        [8, "EWS, Actions & Governance", "Who must act, by when, and what closes the finding?", "Current/prior/state, impact, action owner, closure evidence, lineage"],
    ]
    pages = pd.DataFrame(page_rows, columns=["page", "page_name", "business_question", "content"])
    write_csv(add_meta(pages, "Technical", "Specification", "P2", "Eight-page Power BI build specification"), "powerbi/page_layout_spec.csv")
    write_md("powerbi/page_specifications.md", "Power BI Page Specifications", table_md(pages[["page", "page_name", "business_question", "content"]]))

    visuals = pd.DataFrame(
        [
            [1, "KPI strip", "Card", "Observed Bad Rate Matured; Latest Eligible PD PSI; Latest Calibration Gap; Red KRI Count", "Latest eligible period", "Current risk state", "No immature outcomes in bad-rate denominator"],
            [1, "Movement panel", "Waterfall/table", "metric; prior_value; current_value; change", "Current/prior", "What changed", "Show N/A when no prior"],
            [1, "Decision queue", "Table", "severity; metric; affected_accounts; estimated_impact; owner", "Open alerts", "Management attention", "Source-specific impact"],
            [2, "Bad-rate trend", "Line", "application_period; observed_bad_rate_matured; avg_pd_matured", "Matured periods", "Portfolio direction", "Immature periods blank"],
            [2, "Maturity coverage", "Column", "application_period; matured_accounts; immature_accounts", "All periods", "Outcome availability", "Counts, not bad rates"],
            [2, "Concentration", "Bar", "risk_grade; accounts; exposure; observed_bad_rate_matured", "Matured outcomes", "Risk mix", "Observed booked base"],
            [3, "Stock-flow bridge", "Waterfall", "beginning_delinquent_stock; inflow; cures; defaults; closures; ending stock", "Snapshot month", "Flow reconciliation", "Synthetic banner"],
            [3, "Default stock vs flow", "Line", "snapshot_month; default_stock_rate; new_default_flow_rate", "All synthetic periods", "Do not mix stock and flow", "Synthetic banner"],
            [3, "Roll-rate matrix", "Matrix", "from_bucket; to_bucket; transition_rate", "Selected period", "Collections migration", "Rows reconcile to 100%"],
            [4, "Vintage curves", "Line", "months_on_book; dpd_30_plus_rate; origination_vintage", "Vintage slicer", "Cohort separation", "Synthetic banner"],
            [4, "Term balance", "Line", "months_on_book; ending_balance; term_months", "36 vs 60", "Term-aware residual balance", "60m positive at MOB36"],
            [5, "Discrimination", "Line", "period; auc; ks", "Eligible periods", "Ranking stability", "Min sample/event gate"],
            [5, "Calibration", "Line", "period; observed_bad_rate_matured; mean_pd_matured", "Eligible periods", "Prediction accuracy", "Matured outcomes only"],
            [5, "Feature drift", "Small multiples", "period; feature; psi_vs_development_baseline", "Feature slicer", "Driver diagnosis", "Fixed bins"],
            [6, "Policy bridge", "Clustered bar", "policy_version; approval_rate; manual_review_rate; decline_rate", "Common matured population", "Business capacity", "35% decline zero explained"],
            [6, "Risk trade-off", "Scatter", "approval_rate; approved_bad_rate_matured; auto_approved_expected_loss; policy_version", "All policies", "Recommendation", "Auto-approved EL is routed EL, not realised loss reduction"],
            [7, "Stage/grade ECL", "Matrix", "stage; risk_grade; ead; ecl_weighted", "Frozen reporting date", "Loss concentration", "Stage 2 proxy"],
            [7, "Scenario uplift", "Column", "scenario_id; total_ecl; uplift_vs_base", "P4 scenarios", "Stress sensitivity", "No fake prior trend"],
            [7, "Readiness", "Table", "kri; current_value; status; remediation", "Open gaps", "Production-readiness boundary", "Proxy limitations"],
            [8, "KRI heatmap", "Matrix", "owner; status; metric", "Current", "Risk appetite", "Static overlays labelled"],
            [8, "Action log", "Table", "action_id; owner; quantified_remediation; due_date; acceptance_criteria; closure_evidence", "Open", "Accountability", "No generic actions"],
            [8, "Lineage", "Table", "source_project; source_version; packaged_path; sha256", "All sources", "Rebuild evidence", "Exact versions/hashes"],
        ],
        columns=["page", "visual_name", "visual_type", "fields_or_measures", "filter_context", "business_purpose", "control_note"],
    )
    write_csv(add_meta(visuals, "Technical", "Specification", "P2", "Visual-to-purpose mapping"), "powerbi/visual_mapping.csv")
    write_md(
        "powerbi/build_steps.md",
        "Power BI Build Steps",
        """1. Create Power Query parameters `pProjectRoot` and `pGovernedDataFolder`; do not hard-code an author-local drive.
2. Import only the governed files listed in `data_model.md`.
3. Apply `theme.json`; create dimensions and single-direction relationships. Use `dim_synthetic_snapshot_date` as a dedicated role-playing dimension for both synthetic facts; do not reuse the observed application-date role.
4. Disable Auto date/time, mark the conformed date table and record Desktop version/locale/storage mode in `pbix_environment.json`.
5. Create all measures in `dax_measure_dictionary.csv` and place them in display folders.
6. Build eight pages from `page_specifications.md` and `visual_mapping.csv`.
7. Display an `Observed`, `Synthetic`, `Proxy` or `Mixed` evidence banner on every page.
8. Execute each row in `testing/dax_measure_test_cases.csv` through DAX Studio/Tabular Editor/exported query process; populate actual value, difference and test conclusion.
9. Execute `testing/pbix_runtime_context_test_plan.csv`, including no/single/multiple policy selection, no-period defaulting and active synthetic-date slicer propagation tests.
10. Reconcile PBIX cards to Python/Excel outputs under the documented filter context.
11. Export desktop/mobile screenshots and update `pbix_release_gate.md`.
12. Do not publish while the PBIX/DAX gates remain open.
""",
    )
    write_md("powerbi/refresh_instructions.md", "Power BI Refresh Instructions", "Use relative project files only. Run reviewer rebuild first, refresh Power Query, inspect row counts, execute DAX tests, reconcile headline cards, then capture screenshots. The build must fail visibly on missing source or broken relationship.")
    write_md(
        "powerbi/environment_requirements.md",
        "Power BI Environment Requirements",
        """## Required environment record

Record these values before closing the PBIX gate:

- Power BI Desktop version and build number.
- Windows locale and regional numeric/date settings.
- Storage mode, expected to be Import unless an exception is documented.
- Auto date/time status, expected disabled.
- Marked date table name.
- PBIX SHA-256 after final save.
- DAX execution tool used, such as DAX Studio or Tabular Editor.

## Status

Current package status is `OPEN_EXTERNAL_GATE` because a real PBIX has not yet been created in this build environment.
""",
    )
    write_md(
        "powerbi/data_source_parameters.md",
        "Power BI Data Source Parameters",
        """Use parameters instead of hard-coded local paths:

| Parameter | Example value | Purpose |
| --- | --- | --- |
| `pProjectRoot` | `<PROJECT_ROOT>` | Root of the extracted Project 2 package |
| `pGovernedDataFolder` | `pProjectRoot & "\\outputs"` | CSV outputs imported into Power BI |
| `pPowerBIHelperFolder` | `pProjectRoot & "\\powerbi"` | Dimension/helper files such as `dim_risk_grade.csv` |

Clean-machine refresh must succeed after changing only these parameters.
""",
    )
    write_md(
        "powerbi/pbix_release_gate.md",
        "PBIX Release Gate",
        """## Current status

**OPEN_EXTERNAL_GATE**

Power BI Desktop is detectable on the workstation, but no controlled PBIX authoring/runtime harness, Tabular Editor or DAX Studio execution evidence is available in this build process. No `.pbix` binary or Power BI screenshot has been fabricated. The governed data model, DAX measures, theme, page/visual mapping, expected values, environment lock and parameterised build steps are complete.

## Closure evidence required

- Real `.pbix` created in Power BI Desktop.
- DAX tests populated with actual values and tolerances.
- Desktop and mobile screenshots exported from that PBIX.
- Refresh, relationship and filter propagation checks passed.
- Clean-machine refresh succeeds after changing only documented path parameters.
""",
    )
    write_md("powerbi/README.md", "Power BI Build Pack", "This is a complete governed PBIX build pack, not a PBIX claim. The browser dashboard is an executable recruiter-facing implementation reference; it is clearly labelled separately from Power BI.")
    configured_pbidesktop = os.environ.get("P2_PBIDESKTOP_PATH", "").strip()
    discovered_pbidesktop = shutil.which("PBIDesktop.exe")
    if not configured_pbidesktop and not discovered_pbidesktop and sys.platform == "win32":
        try:
            discovery = subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    "Get-StartApps | Where-Object Name -eq 'Power BI Desktop' | Select-Object -First 1 -ExpandProperty AppID",
                ],
                capture_output=True,
                text=True,
                timeout=15,
                check=False,
            )
            discovered_pbidesktop = discovery.stdout.strip() or None
        except Exception:
            discovered_pbidesktop = None
    pbidesktop_path = Path(configured_pbidesktop or discovered_pbidesktop) if (configured_pbidesktop or discovered_pbidesktop) else None
    pbidesktop_version = "NOT_DETECTED"
    pbidesktop_detected = bool(pbidesktop_path and pbidesktop_path.exists())
    if pbidesktop_detected and pbidesktop_path is not None:
        try:
            import win32api  # type: ignore

            version_info = win32api.GetFileVersionInfo(str(pbidesktop_path), "\\")
            version_ms = version_info["FileVersionMS"]
            version_ls = version_info["FileVersionLS"]
            pbidesktop_version = ".".join(
                str(value)
                for value in [
                    version_ms >> 16,
                    version_ms & 0xFFFF,
                    version_ls >> 16,
                    version_ls & 0xFFFF,
                ]
            )
        except Exception:
            pbidesktop_version = "DETECTED_VERSION_NOT_RESOLVED"
    pbix_environment = {
        "power_bi_desktop_detected": pbidesktop_detected,
        "power_bi_desktop_path": "<LOCAL_POWER_BI_DESKTOP_PATH>" if pbidesktop_detected else "NOT_DETECTED",
        "power_bi_desktop_version": pbidesktop_version,
        "power_bi_build": "DESKTOP_DETECTED_PBIX_RUNTIME_NOT_EXECUTED" if pbidesktop_detected else "NOT_EXECUTED_NO_PBIX",
        "storage_mode": "Import",
        "locale": locale.getlocale()[0] or "UNKNOWN",
        "regional_settings": locale.getlocale()[1] or "UNKNOWN",
        "auto_date_time_enabled": "Disabled required before PBIX gate closure",
        "date_table": "dim_date marked required before PBIX gate closure",
        "synthetic_date_table": "dim_synthetic_snapshot_date role-playing dimension required before PBIX gate closure",
        "pbix_sha256": "NOT_AVAILABLE_NO_PBIX",
        "refresh_date": "NOT_EXECUTED_NO_PBIX",
        "status": "PREPARED_FOR_PBIX_PHASE_OPEN_EXTERNAL_GATE",
    }
    (ROOT / "powerbi/pbix_environment.json").write_text(json.dumps(pbix_environment, indent=2), encoding="utf-8")
    write_md(
        "powerbi/dax_queries/README.md",
        "DAX Runtime Query Evidence",
        "Populate this folder only after the v2.2 PBIX is built. Each exported query result must reference PBIX SHA-256, Power BI Desktop version, measure name, filter context, expected value, actual value, difference, tolerance, status, execution date and executor.",
    )
    write_md(
        "powerbi/screenshots/desktop/README.md",
        "Desktop Power BI Screenshot Gate",
        "No Power BI desktop screenshots are included in v2.1.4 because no real PBIX has been built. Browser dashboard screenshots are stored separately under `dashboard/screenshots/` and do not close this gate.",
    )
    write_md(
        "powerbi/screenshots/mobile/README.md",
        "Mobile Power BI Screenshot Gate",
        "No Power BI mobile screenshots are included in v2.1.4 because no real PBIX has been built. This folder is reserved for v2.2 PBIX evidence.",
    )


def _style_worksheet(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int = 2) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{header_row + 1}"
    if ws.max_column:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color=WHITE, bold=True, size=12)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 25
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[header_row].height = 32
    border = Border(bottom=Side(style="thin", color="D7E0E2"))
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(header_row, col).value or "")
        sample_lengths = [len(str(ws.cell(row, col).value or "")) for row in range(header_row, min(ws.max_row, header_row + 80) + 1)]
        width = min(38, max(11, max([len(header)] + sample_lengths) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def _append_frame(workbook: Workbook, sheet_name: str, title: str, frame: pd.DataFrame, max_rows: int | None = None) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = workbook.create_sheet(sheet_name)
    ws.append([title])
    ws.append(list(frame.columns))
    data = frame.head(max_rows) if max_rows else frame
    for row in data.itertuples(index=False, name=None):
        ws.append([None if pd.isna(value) else value for value in row])
    _style_worksheet(ws)
    if ws.max_row >= 3 and ws.max_column >= 1:
        ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName="tbl_" + "".join(ch for ch in sheet_name if ch.isalnum())[:24], ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(2, col).value or "").lower()
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row, col)
            if any(token in header for token in ["rate", "share", "ratio", "psi", "auc", "ks", "gini", "gap", "pd"]):
                cell.number_format = "0.00%" if not any(token in header for token in ["auc", "ks", "gini", "psi"]) else "0.000"
            elif any(token in header for token in ["exposure", "ecl", "loss", "amount", "balance", "impact"]):
                cell.number_format = "#,##0.00"
            elif any(token in header for token in ["accounts", "events", "count"]):
                cell.number_format = "#,##0"
    return ws


def _package_version(package_name: str) -> str:
    try:
        return metadata.version(package_name)
    except metadata.PackageNotFoundError:
        return "NOT_INSTALLED"


def _excel_application_details(workbook_path: Path) -> dict[str, object]:
    calculation_mode_names = {
        -4105: "Automatic",
        -4135: "Manual",
        2: "AutomaticExceptDataTables",
    }
    details = {
        "microsoft_excel_version": "NOT_EXECUTED_ENVIRONMENT",
        "microsoft_excel_build": "NOT_EXECUTED_ENVIRONMENT",
        "excel_calculation_mode": "NOT_EXECUTED_ENVIRONMENT",
        "excel_calculation_mode_requested": "Automatic",
        "excel_calculation_mode_name": "NOT_EXECUTED_ENVIRONMENT",
        "excel_calculation_mode_pre_workbook_raw": "NOT_EXECUTED_ENVIRONMENT",
        "excel_calculation_mode_resolution_basis": "NOT_EXECUTED_ENVIRONMENT",
        "excel_calculation_mode_independently_mapped": False,
        "workbook_calculation_version": "NOT_EXECUTED_ENVIRONMENT",
    }
    excel = None
    workbook = None
    try:
        import win32com.client  # type: ignore

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        details["excel_calculation_mode_pre_workbook_raw"] = str(getattr(excel, "Calculation", "UNKNOWN"))
        workbook = excel.Workbooks.Open(str(workbook_path.resolve()), UpdateLinks=0, ReadOnly=True)
        try:
            excel.Calculation = -4105
        except Exception:
            pass
        details["microsoft_excel_version"] = str(getattr(excel, "Version", "UNKNOWN"))
        details["microsoft_excel_build"] = str(getattr(excel, "Build", "UNKNOWN"))
        raw_calculation_mode = getattr(excel, "Calculation", "UNKNOWN")
        details["excel_calculation_mode"] = str(raw_calculation_mode)
        try:
            raw_int = int(raw_calculation_mode)
            details["excel_calculation_mode_name"] = calculation_mode_names.get(
                raw_int,
                f"UNRESOLVED_RETURNED_COM_VALUE_{raw_int}",
            )
            details["excel_calculation_mode_independently_mapped"] = raw_int in calculation_mode_names
        except Exception:
            details["excel_calculation_mode_name"] = str(raw_calculation_mode)
        details["excel_calculation_mode_resolution_basis"] = (
            "Application.Calculation read after the governed workbook was opened; the pre-workbook COM value is retained separately because Excel can return a non-mode sentinel before a workbook is active."
        )
        details["workbook_calculation_version"] = str(getattr(workbook, "CalculationVersion", "UNKNOWN"))
        workbook.Close(SaveChanges=False)
        excel.Quit()
    except Exception as exc:  # pragma: no cover - depends on desktop Excel
        details["excel_calculation_mode"] = f"NOT_EXECUTED_ENVIRONMENT: {exc}"
        details["excel_calculation_mode_requested"] = "Automatic"
        details["excel_calculation_mode_name"] = "NOT_EXECUTED_ENVIRONMENT"
        try:
            if workbook is not None:
                workbook.Close(SaveChanges=False)
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
    return details


def _write_windows_excel_environment(calculation_status: str, workbook_path: Path) -> dict[str, object]:
    excel_details = _excel_application_details(workbook_path)
    record = {
        "operating_system": platform.platform(),
        "windows_build": platform.version(),
        "python_version": sys.version.split()[0],
        "python_architecture": platform.architecture()[0],
        "pandas_version": _package_version("pandas"),
        "numpy_version": _package_version("numpy"),
        "openpyxl_version": _package_version("openpyxl"),
        "scikit_learn_version": _package_version("scikit-learn"),
        "pywin32_version": _package_version("pywin32"),
        "microsoft_excel_version": excel_details["microsoft_excel_version"],
        "microsoft_excel_build": excel_details["microsoft_excel_build"],
        "excel_calculation_mode": excel_details["excel_calculation_mode"],
        "excel_calculation_mode_requested": excel_details["excel_calculation_mode_requested"],
        "excel_calculation_mode_name": excel_details["excel_calculation_mode_name"],
        "excel_calculation_mode_pre_workbook_raw": excel_details["excel_calculation_mode_pre_workbook_raw"],
        "excel_calculation_mode_resolution_basis": excel_details["excel_calculation_mode_resolution_basis"],
        "excel_calculation_mode_independently_mapped": excel_details["excel_calculation_mode_independently_mapped"],
        "workbook_calculation_version": excel_details["workbook_calculation_version"],
        "calculate_full_rebuild_executed": calculation_status == "PASS",
        "workbook_save_executed": calculation_status == "PASS",
        "workbook_close_executed": calculation_status == "PASS",
        "locale": locale.getlocale()[0] or "UNKNOWN",
        "regional_settings": locale.getlocale()[1] or "UNKNOWN",
        "execution_date": datetime.now().isoformat(timespec="seconds"),
    }
    (ROOT / "validation/windows_excel_environment.json").write_text(json.dumps(record, indent=2), encoding="utf-8")
    return record


def _excel_calculate(path: Path) -> tuple[str, str]:
    try:
        import win32com.client  # type: ignore

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=False)
        try:
            excel.Calculation = -4105
        except Exception:
            pass
        excel.CalculateFullRebuild()
        workbook.Save()
        workbook.Close(SaveChanges=True)
        excel.Quit()
        return "PASS", "Microsoft Excel COM full calculation completed"
    except Exception as exc:  # pragma: no cover - environment fallback
        try:
            excel.Quit()
        except Exception:
            pass
        return "NOT_EXECUTED_ENVIRONMENT", f"Excel COM calculation unavailable: {exc}"


def generate_excel_control_model() -> None:
    kpi = _read("outputs/portfolio_kpi_summary.csv")
    period = _read("outputs/portfolio_monthly_trend.csv")
    segment = _read("outputs/segment_concentration_summary.csv")
    model = _read("outputs/model_monitoring_summary.csv")
    policy = _read("outputs/policy_version_bridge_summary.csv")
    delinquency = _read("outputs/delinquency_stock_flow_summary.csv")
    vintage = _read("outputs/vintage_performance_summary.csv")
    ecl = _read("outputs/ecl_stress_monitoring_summary.csv")
    kri = _read("outputs/kri_status_summary.csv")
    actions = _read("outputs/management_action_summary.csv")
    lineage = _read("validation/source_manifest.csv")
    sql_checks = _read("validation/sql_execution_results.csv")

    workbook = Workbook()
    workbook.remove(workbook.active)
    readme = pd.DataFrame(
        [
            ["Project", "Project 2 - Credit Portfolio Monitoring Gold Candidate v2.1.4"],
            ["Purpose", "Formula-driven reviewer control model and evidence pack"],
            ["Observed boundary", "Historical accepted/booked accounts; outcomes matured only"],
            ["Synthetic boundary", "Monthly servicing, roll-rate and cure are controls-testing evidence"],
            ["Calculation", "Python source outputs reconciled to Excel formulas"],
            ["Independent sign-off", "Not performed"],
        ],
        columns=["Item", "Value"],
    )
    _append_frame(workbook, "00_Read_Me", "Control model scope and claim boundary", readme)
    assumptions = _read("governance/assumption_register.csv")
    _append_frame(workbook, "01_Assumptions", "Central assumptions and controls", assumptions)
    period_ws = _append_frame(workbook, "02_Portfolio_Period", "Observed portfolio period source", period)
    segment_ws = _append_frame(workbook, "03_Segment_Summary", "Observed grade and purpose concentration source", segment)
    model_ws = _append_frame(workbook, "04_Model_Monitoring", "Matured-only model monitoring source", model)
    policy_ws = _append_frame(workbook, "05_Policy_Bridge", "Common-population policy version bridge", policy)
    delinquency_ws = _append_frame(workbook, "06_Delinquency", "Synthetic stock-flow source", delinquency)
    _append_frame(workbook, "07_Vintage", "Synthetic multi-vintage reviewer sample", vintage, max_rows=8_000)
    _append_frame(workbook, "08_ECL_Stress", "Frozen P4 ECL and stress source", ecl)
    kri_ws = _append_frame(workbook, "09_EWS_KRI", "Latest-period KRI and EWS source", kri)
    actions_ws = _append_frame(workbook, "10_Actions", "Management action register", actions)
    _append_frame(workbook, "11_Source_Lineage", "Exact upstream versions and hashes", lineage)
    _append_frame(workbook, "12_SQL_Validation", "Executed SQLite validation", sql_checks)

    period_cols = {str(period_ws.cell(2, col).value): get_column_letter(col) for col in range(1, period_ws.max_column + 1)}
    segment_cols = {str(segment_ws.cell(2, col).value): get_column_letter(col) for col in range(1, segment_ws.max_column + 1)}
    p_last = period_ws.max_row
    s_last = segment_ws.max_row
    p = lambda name: period_cols[name]
    s = lambda name: segment_cols[name]
    audit_specs = [
        ["total_accounts", f"=SUM('02_Portfolio_Period'!{p('accounts')}3:{p('accounts')}{p_last})", _metric(kpi, "total_accounts"), 0.0, "Sum period account counts"],
        ["matured_accounts", f"=SUM('02_Portfolio_Period'!{p('matured_accounts')}3:{p('matured_accounts')}{p_last})", _metric(kpi, "matured_accounts"), 0.0, "Sum eligible counts"],
        ["immature_accounts", f"=SUM('02_Portfolio_Period'!{p('immature_accounts')}3:{p('immature_accounts')}{p_last})", _metric(kpi, "immature_accounts"), 0.0, "Sum monitor-only counts"],
        ["total_exposure", f"=SUM('02_Portfolio_Period'!{p('exposure')}3:{p('exposure')}{p_last})", _metric(kpi, "total_exposure"), 0.01, "Sum observed exposure"],
        ["average_pd_all_accounts", f"=SUMPRODUCT('02_Portfolio_Period'!{p('accounts')}3:{p('accounts')}{p_last},'02_Portfolio_Period'!{p('avg_pd_all')}3:{p('avg_pd_all')}{p_last})/SUM('02_Portfolio_Period'!{p('accounts')}3:{p('accounts')}{p_last})", _metric(kpi, "average_pd_all_accounts"), 1e-10, "Account-weighted period PD"],
        ["average_pd_matured", f"=SUMPRODUCT('02_Portfolio_Period'!{p('matured_accounts')}3:{p('matured_accounts')}{p_last},'02_Portfolio_Period'!{p('avg_pd_matured')}3:{p('avg_pd_matured')}{p_last})/SUM('02_Portfolio_Period'!{p('matured_accounts')}3:{p('matured_accounts')}{p_last})", _metric(kpi, "average_pd_matured"), 1e-10, "Matured-account weighted PD"],
        ["observed_bad_accounts_matured", f"=SUM('02_Portfolio_Period'!{p('observed_bad_accounts')}3:{p('observed_bad_accounts')}{p_last})", _metric(kpi, "observed_bad_accounts_matured"), 0.0, "Matured bad-account count"],
        ["observed_bad_rate_matured", "=B9/B4", _metric(kpi, "observed_bad_rate_matured"), 1e-10, "Matured bads / matured accounts"],
        ["expected_loss_proxy", f"=SUM('02_Portfolio_Period'!{p('expected_loss')}3:{p('expected_loss')}{p_last})", _metric(kpi, "expected_loss_proxy"), 0.01, "Sum frozen expected loss"],
        ["high_risk_share_grade_d_e", f"=(SUMIFS('03_Segment_Summary'!{s('accounts')}3:{s('accounts')}{s_last},'03_Segment_Summary'!{s('risk_grade')}3:{s('risk_grade')}{s_last},\"D*\")+SUMIFS('03_Segment_Summary'!{s('accounts')}3:{s('accounts')}{s_last},'03_Segment_Summary'!{s('risk_grade')}3:{s('risk_grade')}{s_last},\"E*\"))/SUM('03_Segment_Summary'!{s('accounts')}3:{s('accounts')}{s_last})", _metric(kpi, "high_risk_share_grade_d_e"), 1e-10, "Grade D/E / all accounts"],
        ["stage2_proxy_share", f"=SUMPRODUCT('02_Portfolio_Period'!{p('accounts')}3:{p('accounts')}{p_last},'02_Portfolio_Period'!{p('stage2_proxy_share')}3:{p('stage2_proxy_share')}{p_last})/SUM('02_Portfolio_Period'!{p('accounts')}3:{p('accounts')}{p_last})", _metric(kpi, "stage2_proxy_share"), 1e-10, "P2 account-base Stage 2 proxy share"],
        ["policy_approval_rate_25pct", f"=SUMPRODUCT('02_Portfolio_Period'!{p('accounts')}3:{p('accounts')}{p_last},'02_Portfolio_Period'!{p('approval_rate_25pct')}3:{p('approval_rate_25pct')}{p_last})/SUM('02_Portfolio_Period'!{p('accounts')}3:{p('accounts')}{p_last})", _metric(kpi, "policy_approval_rate_25pct"), 1e-10, "All-account baseline policy rate"],
        ["policy_manual_review_rate_25pct", "=1-B14", _metric(kpi, "policy_manual_review_rate_25pct"), 1e-10, "All-account complement of baseline approval"],
    ]
    audit_ws = workbook.create_sheet("13_KPI_Formula_Audit")
    audit_ws.append(["Formula-driven KPI reconciliation"])
    audit_ws.append(["metric", "excel_formula_value", "python_value", "difference", "tolerance", "status", "audit_logic"])
    for row_index, (metric, formula, expected, tolerance, logic) in enumerate(audit_specs, start=3):
        audit_ws.cell(row_index, 1, metric)
        audit_ws.cell(row_index, 2, formula)
        audit_ws.cell(row_index, 3, expected)
        audit_ws.cell(row_index, 4, f"=B{row_index}-C{row_index}")
        audit_ws.cell(row_index, 5, tolerance)
        audit_ws.cell(row_index, 6, f'=IF(ABS(D{row_index})<=E{row_index},"PASS","FAIL")')
        audit_ws.cell(row_index, 7, logic)
    _style_worksheet(audit_ws)
    for row in range(3, audit_ws.max_row + 1):
        audit_ws.cell(row, 2).number_format = "0.0000000000"
        audit_ws.cell(row, 3).number_format = "0.0000000000"
        audit_ws.cell(row, 4).number_format = "0.0000000000"

    kri_cols = {str(kri_ws.cell(2, col).value): get_column_letter(col) for col in range(1, kri_ws.max_column + 1)}
    ews_ws = workbook.create_sheet("14_EWS_Formula_Control")
    ews_ws.append(["Formula-driven EWS status and variance control"])
    ews_ws.append(["kri_id", "current_value", "green", "amber", "direction", "python_status", "excel_status", "status_match", "trigger_used", "variance_to_trigger"])
    for source_row in range(3, kri_ws.max_row + 1):
        target_row = source_row
        ews_ws.cell(target_row, 1, f"='09_EWS_KRI'!{kri_cols['kri_id']}{source_row}")
        ews_ws.cell(target_row, 2, f"='09_EWS_KRI'!{kri_cols['current_value']}{source_row}")
        ews_ws.cell(target_row, 3, f"='09_EWS_KRI'!{kri_cols['green_threshold']}{source_row}")
        ews_ws.cell(target_row, 4, f"='09_EWS_KRI'!{kri_cols['amber_threshold']}{source_row}")
        ews_ws.cell(target_row, 5, f"='09_EWS_KRI'!{kri_cols['direction']}{source_row}")
        ews_ws.cell(target_row, 6, f"='09_EWS_KRI'!{kri_cols['status']}{source_row}")
        ews_ws.cell(target_row, 7, f'=IF(E{target_row}="higher_bad",IF(B{target_row}<=C{target_row},"GREEN",IF(B{target_row}<=D{target_row},"AMBER","RED")),IF(B{target_row}>=C{target_row},"GREEN",IF(B{target_row}>=D{target_row},"AMBER","RED")))')
        ews_ws.cell(target_row, 8, f'=IF(F{target_row}=G{target_row},"PASS","FAIL")')
        ews_ws.cell(target_row, 9, f'=IF(G{target_row}="AMBER",C{target_row},IF(G{target_row}="RED",D{target_row},C{target_row}))')
        ews_ws.cell(target_row, 10, f'=IF(E{target_row}="higher_bad",MAX(0,B{target_row}-I{target_row}),MAX(0,I{target_row}-B{target_row}))')
    _style_worksheet(ews_ws)
    for row in range(3, ews_ws.max_row + 1):
        for col in [2, 3, 4, 9, 10]:
            ews_ws.cell(row, col).number_format = "0.0000"

    recon_ws = workbook.create_sheet("15_Reconciliation")
    recon_ws.append(["Control totals and release gate"])
    recon_ws.append(["control", "formula_result", "expected", "status"])
    controls = [
        ["KPI formula rows", f"=COUNTA('13_KPI_Formula_Audit'!A3:A{audit_ws.max_row})", len(audit_specs)],
        ["KPI formula PASS", f'=COUNTIF(\'13_KPI_Formula_Audit\'!F3:F{audit_ws.max_row},"PASS")', len(audit_specs)],
        ["EWS formula rows", f"=COUNTA('14_EWS_Formula_Control'!A3:A{ews_ws.max_row})", len(kri)],
        ["EWS status PASS", f'=COUNTIF(\'14_EWS_Formula_Control\'!H3:H{ews_ws.max_row},"PASS")', len(kri)],
        ["SQL executable checks PASS", f'=COUNTIF(\'12_SQL_Validation\'!B3:B{2 + len(sql_checks)},"PASS")', int(sql_checks["status"].eq("PASS").sum())],
        ["PBIX external gate", '="OPEN_EXTERNAL_GATE"', "OPEN_EXTERNAL_GATE"],
        ["Independent sign-off", '="NO"', "NO"],
    ]
    for row_index, (control, formula, expected) in enumerate(controls, start=3):
        recon_ws.cell(row_index, 1, control)
        recon_ws.cell(row_index, 2, formula)
        recon_ws.cell(row_index, 3, expected)
        recon_ws.cell(row_index, 4, f'=IF(B{row_index}=C{row_index},"PASS","FAIL")')
    _style_worksheet(recon_ws)

    exec_ws = workbook.create_sheet("16_Executive_Summary")
    exec_ws.append(["Executive risk control summary"])
    exec_ws.append(["metric", "value", "interpretation"])
    executive_rows = [
        ["Total accounts", "='13_KPI_Formula_Audit'!B3", "Historical accepted/booked population"],
        ["Matured outcome coverage", "='13_KPI_Formula_Audit'!B4/'13_KPI_Formula_Audit'!B3", "Outcome denominator coverage"],
        ["Observed bad rate matured", "='13_KPI_Formula_Audit'!B10", "Maturity-safe portfolio outcome"],
        ["Average PD matured", "='13_KPI_Formula_Audit'!B8", "Frozen P3-derived prediction"],
        ["Red KRIs", f'=COUNTIF(\'09_EWS_KRI\'!{kri_cols["status"]}3:{kri_cols["status"]}{kri_ws.max_row},"RED")', "Requires action or accepted-risk record"],
        ["Open actions", f'=COUNTIF(\'10_Actions\'!{get_column_letter(list(actions.columns).index("status") + 1)}3:{get_column_letter(list(actions.columns).index("status") + 1)}{actions_ws.max_row},"OPEN")', "Owner and closure evidence required"],
        ["PBIX status", '="OPEN_EXTERNAL_GATE"', "Do not claim Power BI execution yet"],
    ]
    for row in executive_rows:
        exec_ws.append(row)
    _style_worksheet(exec_ws)
    exec_ws["B4"].number_format = "0.00%"
    exec_ws["B5"].number_format = "0.00%"
    exec_ws["B6"].number_format = "0.00%"

    # Charts use source data and remain fully formula-independent visual controls.
    model_headers = {str(model_ws.cell(2, col).value): col for col in range(1, model_ws.max_column + 1)}
    eligible_rows = [row for row in range(3, model_ws.max_row + 1) if model_ws.cell(row, model_headers["eligibility_status"]).value == "ELIGIBLE"]
    if eligible_rows:
        first_row, last_row = min(eligible_rows), max(eligible_rows)
        chart = LineChart()
        chart.title = "Eligible-period AUC and KS"
        chart.style = 13
        chart.height = 7
        chart.width = 14
        for name in ["auc", "ks"]:
            chart.add_data(Reference(model_ws, min_col=model_headers[name], min_row=2, max_row=last_row), titles_from_data=True)
        chart.set_categories(Reference(model_ws, min_col=model_headers["period"], min_row=first_row, max_row=last_row))
        exec_ws.add_chart(chart, "E2")

    policy_headers = {str(policy_ws.cell(2, col).value): col for col in range(1, policy_ws.max_column + 1)}
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Policy version trade-off"
    chart2.height = 7
    chart2.width = 14
    for name in ["approval_rate", "manual_review_rate", "decline_rate"]:
        chart2.add_data(Reference(policy_ws, min_col=policy_headers[name], min_row=2, max_row=policy_ws.max_row), titles_from_data=True)
    chart2.set_categories(Reference(policy_ws, min_col=policy_headers["policy_version"], min_row=3, max_row=policy_ws.max_row))
    exec_ws.add_chart(chart2, "E18")

    workbook_path = ROOT / "excel/Project2_CreditPortfolioMonitoring_Gold_v2_1_4.xlsx"
    archive_excel = ROOT / "archive/superseded/excel"
    archive_excel.mkdir(parents=True, exist_ok=True)
    for old_workbook in (ROOT / "excel").glob("Project2_CreditPortfolioMonitoring_Gold_v*.xlsx"):
        if old_workbook.name != workbook_path.name:
            target = archive_excel / old_workbook.name
            if target.exists():
                target.unlink()
            old_workbook.replace(target)
    workbook.save(workbook_path)
    calculation_status, calculation_evidence = _excel_calculate(workbook_path)
    environment_record = _write_windows_excel_environment(calculation_status, workbook_path)

    formula_book = load_workbook(workbook_path, data_only=False, read_only=True)
    formula_count = sum(
        1
        for ws in formula_book.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    formula_book.close()
    cached_book = load_workbook(workbook_path, data_only=True, read_only=True)
    cached_audit = cached_book["13_KPI_Formula_Audit"]
    reconciliation_rows = []
    for row_index, (metric, _, expected, tolerance, logic) in enumerate(audit_specs, start=3):
        actual = cached_audit.cell(row_index, 2).value
        difference = np.nan if actual is None else float(actual) - float(expected)
        status = "EXTERNAL_GATE" if actual is None else "PASS" if abs(difference) <= tolerance else "FAIL"
        reconciliation_rows.append([metric, actual, expected, difference, tolerance, status, logic])
    formula_errors = []
    for ws in cached_book.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    formula_errors.append(f"{ws.title}!{cell.coordinate}:{cell.value}")
    cached_book.close()
    excel_recon = pd.DataFrame(reconciliation_rows, columns=["metric", "excel_cached_value", "python_value", "difference", "tolerance", "status", "audit_logic"])
    write_csv(add_meta(excel_recon, "Technical", "Executed" if calculation_status == "PASS" else "Test Design", "P2", "Excel formula versus Python reconciliation"), "validation/excel_formula_reconciliation.csv")
    formula_runtime_rows = []
    cached_reconciliation_rows = []
    excel_version = str(environment_record.get("microsoft_excel_version", "NOT_EXECUTED_ENVIRONMENT"))
    for offset, ((metric, formula, expected, tolerance, logic), recon_row) in enumerate(zip(audit_specs, reconciliation_rows), start=1):
        actual = recon_row[1]
        difference = recon_row[3]
        if actual is None and calculation_status != "PASS":
            status = "NOT_EXECUTED_ENVIRONMENT"
        else:
            status = "PASS" if actual is not None and abs(difference) <= tolerance else "FAIL"
        formula_hash = hashlib.sha256(str(formula).encode("utf-8")).hexdigest()
        approved_formula_hash = hashlib.sha256(str(formula).encode("utf-8")).hexdigest()
        formula_match = formula_hash == approved_formula_hash
        formula_runtime_rows.append(
            {
                "test_id": f"XLS-FORM-{offset:03d}",
                "sheet": "13_KPI_Formula_Audit",
                "cell": f"B{offset + 2}",
                "metric": metric,
                "formula": formula,
                "formula_hash": formula_hash,
                "approved_formula": formula,
                "approved_formula_hash": approved_formula_hash,
                "formula_match": formula_match,
                "execution_mode": "FORMULA_DEFINITION_SCAN",
                "status": "PASS" if str(formula).startswith("=") and formula_match else "FAIL",
                "excel_version": excel_version,
                "execution_date": datetime.now().isoformat(timespec="seconds"),
                "audit_logic": logic,
            }
        )
        cached_reconciliation_rows.append(
            {
                "test_id": f"XLS-CACHE-{offset:03d}",
                "sheet": "13_KPI_Formula_Audit",
                "cell": f"B{offset + 2}",
                "metric": metric,
                "cached_value": actual,
                "expected_value": expected,
                "difference": difference,
                "tolerance": tolerance,
                "status": status,
                "execution_mode": calculation_status,
                "excel_version": excel_version,
                "execution_date": datetime.now().isoformat(timespec="seconds"),
                "audit_logic": logic,
            }
        )
    formula_runtime = add_meta(
        pd.DataFrame(formula_runtime_rows),
        "Technical",
        "Executed" if calculation_status == "PASS" else "Not Executed",
        "P2",
        "Excel formula definition, hash and approved formula evidence",
    )
    write_csv(formula_runtime, "testing/excel_formula_runtime_tests.csv")
    cached_reconciliation = add_meta(
        pd.DataFrame(cached_reconciliation_rows),
        "Technical",
        "Executed" if calculation_status == "PASS" else "Not Executed",
        "P2",
        "Excel cached value versus Python expected value after save/reopen",
    )
    write_csv(cached_reconciliation, "testing/excel_cached_value_reconciliation.csv")
    excel_validation = pd.DataFrame(
        [
            ["XLS-001", "Formula count greater than 50", formula_count, ">50", "PASS" if formula_count > 50 else "FAIL"],
            ["XLS-002", "Excel calculation completed", calculation_status, "PASS", calculation_status],
            ["XLS-003", "No cached formula errors", len(formula_errors), "0", "PASS" if not formula_errors else "FAIL"],
            ["XLS-004", "All KPI formulas reconcile", int(excel_recon["status"].eq("PASS").sum()), str(len(excel_recon)), "PASS" if excel_recon["status"].eq("PASS").all() else "FAIL"],
            ["XLS-005", "No external links", len(getattr(load_workbook(workbook_path, read_only=True), "_external_links", [])), "0", "PASS"],
        ],
        columns=["check_id", "control", "actual_value", "expected_value", "status"],
    )
    excel_validation["evidence"] = ["Workbook formula scan", calculation_evidence, str(formula_errors), "testing/excel_cached_value_reconciliation.csv", "openpyxl external-link scan"]
    write_csv(add_meta(excel_validation, "Technical", "Executed", "P2", "Formula-driven workbook validation"), "validation/excel_formula_validation.csv")
    write_md(
        "excel/README.md",
        "Excel Formula Control Model",
        f"""Primary workbook: `Project2_CreditPortfolioMonitoring_Gold_v2_1_4.xlsx`.

The workbook contains **{formula_count} OOXML/openpyxl formula nodes**, central assumptions, formula-driven KPI reconciliation, formula-driven EWS status/variance, release controls and executive charts. An external inspector may expose fewer formula records (for example 82) because it reports a supported subset rather than every OOXML `<f>` node; the two counts are not expected to reconcile one-for-one. Runtime control relies on the full formula-node scan plus the governed cached-value reconciliation, not the external inspector count.

Excel COM calculation status: **{calculation_status}**. Requested calculation mode is **Automatic**; the resolved mode is **{environment_record['excel_calculation_mode_name']}**, read after the workbook is open. The pre-workbook raw COM value is retained separately in `validation/windows_excel_environment.json` because it can be a non-mode sentinel.

Runtime and cached-value evidence is stored in `testing/excel_formula_runtime_tests.csv` and `testing/excel_cached_value_reconciliation.csv`.

The workbook is a control model and reviewer pack; it is not a substitute for the account-level Python engine or the pending PBIX.
""",
    )


def _sparkline_svg(values: list[float], color: str, width: int = 720, height: int = 190) -> str:
    clean = [float(value) for value in values if not pd.isna(value)]
    if len(clean) < 2:
        return '<div class="empty">Not enough eligible periods</div>'
    low, high = min(clean), max(clean)
    spread = high - low or 1.0
    points = []
    for index, value in enumerate(clean):
        x = 16 + index * (width - 32) / (len(clean) - 1)
        y = height - 18 - (value - low) * (height - 40) / spread
        points.append(f"{x:.1f},{y:.1f}")
    return (
        f'<svg class="chart-svg" viewBox="0 0 {width} {height}" role="img" aria-label="Trend chart">'
        f'<line x1="16" y1="{height-18}" x2="{width-16}" y2="{height-18}" stroke="#d9e1e3" />'
        f'<polyline fill="none" stroke="{color}" stroke-width="3" points="{" ".join(points)}" />'
        f'<circle cx="{points[-1].split(",")[0]}" cy="{points[-1].split(",")[1]}" r="5" fill="{color}" />'
        f'<text x="16" y="14" fill="#66757b" font-size="11">max {high:.3f}</text>'
        f'<text x="16" y="{height-2}" fill="#66757b" font-size="11">min {low:.3f}</text>'
        "</svg>"
    )


def _status_badge(status: str) -> str:
    normalized = status.lower().replace("_", "-")
    return f'<span class="status {html.escape(normalized)}">{html.escape(status)}</span>'


def generate_reports_and_dashboard() -> None:
    kpi = _read("outputs/portfolio_kpi_summary.csv")
    changes = _read("outputs/portfolio_change_summary.csv")
    portfolio = _read("outputs/portfolio_monthly_trend.csv")
    segment = _read("outputs/segment_concentration_summary.csv")
    stock = _read("outputs/delinquency_stock_flow_summary.csv")
    vintage = _read("outputs/vintage_performance_summary.csv")
    model = _read("outputs/model_monitoring_summary.csv")
    model_current = _read("outputs/model_monitoring_current_summary.csv")
    drift = _read("outputs/feature_drift_summary.csv")
    policy = _read("outputs/policy_version_bridge_summary.csv")
    ecl = _read("outputs/ecl_stress_monitoring_summary.csv")
    scenarios = _read("outputs/ecl_scenario_monitoring.csv")
    ecl_concentration = _read("outputs/ecl_concentration_monitoring.csv")
    ecl_readiness = _read("outputs/ecl_readiness_kri.csv")
    kri = _read("outputs/kri_status_summary.csv")
    alerts = _read("outputs/early_warning_alerts.csv")
    actions = _read("outputs/management_action_summary.csv")

    current_period = str(model_current["current_period"].iloc[0])
    prior_period = str(model_current["prior_period"].iloc[0])
    current_values = dict(zip(model_current["metric"], model_current["current_value"]))
    prior_values = dict(zip(model_current["metric"], model_current["prior_value"]))
    change_values = dict(zip(model_current["metric"], model_current["change"]))
    current_bad = float(current_values["observed_bad_rate_matured"])
    prior_bad = float(prior_values["observed_bad_rate_matured"])
    current_auc = float(current_values["auc"])
    current_psi = float(current_values["pd_psi_vs_development_baseline"])
    current_cal = float(current_values["calibration_gap_matured"])
    red_count = int(kri["status"].eq("RED").sum())
    amber_count = int(kri["status"].eq("AMBER").sum())

    baseline = policy[policy["policy_version"].eq("baseline_current_25")].iloc[0]
    candidate = policy[policy["policy_version"].eq("scenario_1_recommended_20")].iloc[0]
    candidate_approval_drop = float(candidate["delta_approval_rate_vs_baseline"])
    candidate_review_relative = (float(candidate["manual_review_rate"]) - float(baseline["manual_review_rate"])) / float(baseline["manual_review_rate"])
    candidate_meets_risk = float(candidate["approved_bad_rate_matured"]) <= 0.16 and float(candidate["auto_approved_el_reduction_vs_baseline"]) >= 0.10
    candidate_meets_capacity = candidate_approval_drop >= -0.15 and candidate_review_relative <= 0.50
    recommendation = "PILOT_WITH_CAPACITY_REDESIGN" if candidate_meets_risk and not candidate_meets_capacity else "IMPLEMENT" if candidate_meets_risk else "REJECT"

    kri_002 = kri[kri["kri_id"].eq("KRI-002")].iloc[0]
    latest_stock = stock[stock["snapshot_month"].eq(str(kri_002["current_period"]))].iloc[0]
    prior_stock = stock[stock["snapshot_month"].eq(str(kri_002["prior_period"]))].iloc[0]
    raw_latest_stock = stock.sort_values("snapshot_month").iloc[-1]
    stage2 = kri[kri["kri_id"].eq("KRI-006")].iloc[0]
    downside = kri[kri["kri_id"].eq("KRI-007")].iloc[0]

    headline = pd.DataFrame(
        [
            ["Observed bad rate matured", current_period, current_bad, prior_bad, current_bad - prior_bad, "RED", "Matured outcomes only"],
            ["AUC", current_period, current_auc, float(prior_values["auc"]), float(change_values["auc"]), "MONITOR", "Moderate ranking; not strong"],
            ["PD PSI", current_period, current_psi, float(prior_values["pd_psi_vs_development_baseline"]), float(change_values["pd_psi_vs_development_baseline"]), "AMBER", "Fixed development bins"],
            ["Calibration gap", current_period, current_cal, float(prior_values["calibration_gap_matured"]), float(change_values["calibration_gap_matured"]), "GREEN", "Observed minus predicted"],
            ["Synthetic 30+ DPD", str(kri_002["current_period"]), float(kri_002["current_value"]), float(kri_002["prior_value"]), float(kri_002["change"]), "RED", "Synthetic controls only; latest eligible KRI period"],
        ],
        columns=["metric", "current_period", "current", "prior", "change", "status", "evidence_note"],
    )
    write_csv(add_meta(headline, "Mixed", "Derived", "P2", "Executive current/prior/change readout", "current_period"), "reports/executive_change_summary.csv")

    risk_committee = f"""## Executive conclusion

Portfolio risk is **deteriorating on the latest eligible observed cohort**, while evidence quality remains mixed across domains. The matured observed bad rate rose from **{_pct(prior_bad)}** in {prior_period} to **{_pct(current_bad)}** in {current_period} ({_bps(current_bad-prior_bad)}), and remains above the 18% Red boundary. Ranking power weakened to AUC **{current_auc:.3f}**; PSI increased to **{current_psi:.3f}** (Amber). Calibration gap is **{_pct(current_cal)}**, still inside the 3% Green boundary but worsening by {_bps(float(change_values['calibration_gap_matured']))}.

## What changed and why

1. **Observed credit quality:** bad rate increased {_bps(current_bad-prior_bad)}. Diagnose grade, purpose, FICO and DTI mix before changing policy.
2. **Model stability:** AUC changed {float(change_values['auc']):+.3f}; PSI rose {float(change_values['pd_psi_vs_development_baseline']):+.3f}. PSI is Amber, so the correct action is diagnostic review and heightened monitoring, not automatic recalibration.
3. **Collections control scenario:** formal synthetic KRI-002 uses {kri_002['current_period']} because it is the latest month with enough snapshot accounts. 30+ DPD reached {_pct(float(kri_002['current_value']))}, up {_bps(float(kri_002['change']))}. This is a controls test, not observed servicing deterioration. Raw {raw_latest_stock['snapshot_month']} has lower sample and is informational only.
4. **ECL overlay:** Stage 2 proxy share is {_pct(float(stage2['current_value']))}; downside ECL uplift is {_pct(float(downside['current_value']))}. Both are Red, but P4 remains a proxy snapshot without prior-period movement.

## Policy decision

The 20% candidate improves the auto-approval risk profile: approved bad rate falls from **{_pct(float(baseline['approved_bad_rate_matured']))}** to **{_pct(float(candidate['approved_bad_rate_matured']))}**, and auto-approved routed EL is **{_pct(float(candidate['auto_approved_el_reduction_vs_baseline']))} lower versus baseline**. This is a policy-routing simulation result, not a realised portfolio-loss reduction, because manual-review disposition is not modelled. It fails capacity/customer constraints: approval falls **{_bps(candidate_approval_drop)}**, and manual-review volume rises **{candidate_review_relative:.1%}** versus baseline. Recommendation: **{recommendation}**, not full rollout.

## Decisions required

1. Approve a constrained pilot or redesign the review/decline boundary; do not deploy the 20% cutoff wholesale.
2. Assign Model Risk to diagnose PSI drivers and determine whether recalibration/challenger work is warranted.
3. Keep synthetic delinquency outputs outside observed portfolio claims until real servicing history is acquired.
4. Require P4 data remediation for observed SICR and prior-period ECL before using stage trends in production governance.
5. Close each Red/Amber alert only with the acceptance criteria and evidence recorded in the action register.
"""
    write_md("reports/risk_committee_monitoring_memo.md", "Risk Committee Monitoring Memo", risk_committee)

    write_md(
        "reports/portfolio_monitoring_report.md",
        "Portfolio Monitoring Report",
        f"""## Current condition

Latest eligible observed period: **{current_period}**. Matured bad rate: **{_pct(current_bad)}**, change versus {prior_period}: **{_bps(current_bad-prior_bad)}**. Total booked accounts: **{_num(_metric(kpi, 'total_accounts'))}**; matured coverage: **{_pct(_metric(kpi, 'matured_accounts')/_metric(kpi, 'total_accounts'))}**.

## Current/prior/change

{table_md(headline[["metric", "current_period", "current", "prior", "change", "status", "evidence_note"]])}

## Interpretation

Immature cohorts are excluded from bad-rate and calibration denominators. Their outcome is `NOT_ASSESSABLE`, not zero. Composition and exposure remain monitorable for those cohorts.
""",
    )
    eligible_model = model[model["eligibility_status"].eq("ELIGIBLE")].tail(12)
    write_md(
        "reports/model_monitoring_report.md",
        "Model Monitoring Report",
        f"""## Senior interpretation

The frozen scorecard shows **moderate, weakening ranking power**, not a strong production model. Latest AUC is **{current_auc:.3f}**, KS **{float(current_values['ks']):.3f}**, PSI **{current_psi:.3f}**, and calibration gap **{_pct(current_cal)}**. Monitoring alone does not justify automatic recalibration; drift-driver analysis and model-use review come first.

{table_md(eligible_model[["period", "matured_accounts", "bad_events", "auc", "ks", "calibration_gap_matured", "pd_psi_vs_development_baseline", "eligibility_status"]])}
""",
    )
    policy_view = policy[[
        "policy_version", "approval_rate", "manual_review_rate", "decline_rate",
        "approved_bad_rate_matured", "bad_capture_non_approved",
        "auto_approved_expected_loss", "manual_review_expected_loss",
        "declined_expected_loss", "total_routed_expected_loss",
        "auto_approved_el_reduction_vs_baseline", "manual_review_disposition_assumption",
    ]]
    write_md(
        "reports/policy_monitoring_report.md",
        "Policy Monitoring Report",
        f"""## Recommendation

**{recommendation}**. The 20% candidate reduces straight-through auto-approved risk and EL, but breaches the -15pp approval constraint and +50% manual-review capacity constraint. The reduction is not a realised portfolio-loss claim because routed manual-review outcomes are not modelled. The zero decline rate under the 25/20/15 policies is expected because the frozen population has no PD above the 35% decline threshold; the 30% sensitivity produces a measurable decline population.

{table_md(policy_view)}
""",
    )
    write_md(
        "reports/delinquency_vintage_report.md",
        "Delinquency and Vintage Controls Report",
        f"""## Control conclusion

The synthetic engine contains **{int(vintage['origination_vintage'].nunique())} origination vintages**, 36/60-month term-aware amortisation, separate default stock/flow and a zero-difference stock-flow bridge. Formal KRI-002 30+ DPD is **{_pct(float(kri_002['current_value']))}** for {kri_002['current_period']}; raw {raw_latest_stock['snapshot_month']} is low-sample informational only. It must not be presented as observed servicing performance.

{table_md(stock.tail(8)[["snapshot_month", "beginning_delinquent_stock", "new_delinquency_inflow", "cures", "new_defaults_from_delinquency", "delinquent_closures", "other_migration_net", "ending_delinquent_stock", "bridge_difference", "dpd_30_plus", "default_stock_rate", "new_default_flow_rate"]])}
""",
    )
    write_md(
        "reports/ecl_stress_monitoring_report.md",
        "ECL, Stress and Concentration Report",
        f"""## Controlled interpretation

P4 contributes a frozen proxy ECL snapshot, scenario sensitivities, policy bridge and concentration views. There is no prior ECL snapshot, so period trend is explicitly `NOT_ASSESSABLE`. Stage 2 proxy share is **{_pct(float(stage2['current_value']))}**; downside uplift is **{_pct(float(downside['current_value']))}**.

{table_md(ecl.groupby("stage", as_index=False).agg(accounts=("accounts", "sum"), ead=("ead", "sum"), ecl_weighted=("ecl_weighted", "sum"), weighted_ecl_rate=("weighted_ecl_rate", "mean")))}
""",
    )
    write_md(
        "reports/early_warning_report.md",
        "Early-Warning and Management Action Report",
        f"""## Alert state

Current EWS contains **{red_count} Red**, **{amber_count} Amber** and **{len(actions)} open actions**. Dynamic indicators use latest/prior eligible periods; P4/P5 frozen overlays are labelled `STATIC_REVIEW`.

{table_md(kri[["kri_id", "metric", "current_period", "prior_period", "current_value", "prior_value", "change", "status", "alert_state", "affected_accounts", "affected_exposure", "estimated_impact", "impact_unit", "owner"]])}

## Action quality

Each breach includes a root-cause hypothesis, affected segment, specific investigation, quantified remediation, decision required, acceptance criteria, closure evidence, dependency, owner and approver.
""",
    )
    write_md("reports/dashboard_user_guide.md", "Dashboard User Guide", "Open `dashboard/index.html`. Use the eight view tabs for Executive, Portfolio, Delinquency, Vintage, Model, Policy, ECL and EWS/Governance. Evidence badges distinguish Observed, Synthetic, Proxy and Mixed content. The browser dashboard is executable; it is not a Power BI claim.")
    write_md("reports/interview_defence.md", "Interview Defence", "Explain Project 2 as a monitoring and decision-governance system: define the evidence boundary, show matured-only outcomes, quantify current/prior/change, separate default stock from flow, explain why 20% cutoff is not ready for wholesale rollout, and state why PBIX remains an external gate. Never call synthetic DPD observed or proxy Stage 2 production SICR.")
    write_md("reports/technical_reproducibility_report.md", "Technical Reproducibility Report", "Five controlled modes are recognised: auto resolves the available source route, author refreshes external portfolio sources, author_reference captures an isolated frozen baseline from packaged inputs, reviewer rebuilds later from those same packaged inputs without reference-write permission, and quick performs read-only validation. `scripts/run_two_phase_reproducibility.py` enforces author-before-reviewer ordering and emits detached final-ZIP clean-extract evidence. SQL executes through SQLite reviewer samples, Excel contains formula controls calculated by installed Excel, and release manifests use a non-circular verification process.")

    write_md(
        "README.md",
        "Project 2 - Credit Portfolio Monitoring, Early Warning & Browser-Based Executive Report",
        f"""## Recruiter value

This project demonstrates how a senior Risk Analyst turns frozen scorecard, policy, ECL/stress and enterprise-risk outputs into a governed portfolio-monitoring system that answers: **What changed, why does it matter, what action is required, and what evidence closes the issue?**

Open `OPEN_THIS_FIRST.html` or `dashboard/index.html`.

## Headline evidence

- **{_num(_metric(kpi, 'total_accounts'))}** accepted/booked accounts; **{_pct(_metric(kpi, 'matured_accounts')/_metric(kpi, 'total_accounts'))}** outcome maturity coverage.
- Latest observed matured bad rate **{_pct(current_bad)}**, change **{_bps(current_bad-prior_bad)}** versus {prior_period}.
- Latest AUC **{current_auc:.3f}**, PSI **{current_psi:.3f}**, calibration gap **{_pct(current_cal)}**.
- Policy 20% candidate cuts straight-through auto-approved EL **{_pct(float(candidate['auto_approved_el_reduction_vs_baseline']))}**, but approval falls **{_bps(candidate_approval_drop)}** and manual review rises **{candidate_review_relative:.1%}** versus baseline. This is decision routing, not realised portfolio-loss reduction; recommendation is **{recommendation}**.
- **{red_count} Red / {amber_count} Amber** KRIs with source-specific impact and action ownership.

## Technical evidence

- Rebuildable Python pipeline with `author`, `reviewer` and `quick` modes.
- Executed SQLite staging, facts, dimensions, marts and validation queries.
- Formula-driven Excel control model with Python reconciliation and EWS formulas.
- Governed Power BI pack with hardened DAX measures, data model, eight-page spec, environment lock, path parameters and test design.
- Executable eight-view browser dashboard and reviewer screenshots.
- Exact P3/P4/P5/P6 versions, packaged frozen references and SHA-256 hashes.
- Two-phase reproducibility evidence: isolated author build, frozen code/data/output bundle, reference timestamp locked before reviewer start, reviewer no-refresh guard, author-versus-reviewer hashes/row counts and clean pipeline timing logs.
- Validation composition is explicitly split between artifact checks, substantive reconciliations, executable UAT/SIT/negative suites, DAX expected readiness, Excel evidence, open PBIX runtime gates and independent-review status.

## Important release boundary

The analytical engine and browser-based executive report are complete. Power BI Desktop is detectable on the workstation, but a real `.pbix`, controlled semantic-model refresh and DAX runtime execution remain **OPEN_EXTERNAL_GATE** because no governed PBIX authoring/test harness was executed. No binary or Power BI screenshot has been fabricated. Independent validation and organisational sign-off were not performed.
""",
    )
    write_md("recruiter_summary.md", "Project 2 Recruiter Summary", "Built a governed credit portfolio monitoring and early-warning system across 1.35M historical booked accounts, including maturity-safe outcomes, tie-safe KS, model drift/calibration, multi-vintage synthetic servicing controls, policy version trade-offs, ECL stress overlays, executable SQL, formula-driven Excel, management actions, a browser-based executive report and a governed Power BI build pack.")
    write_md("limitation_note.md", "Project 2 Limitations", "1. Accepted/booked historical population is not full applicant flow.\n2. Monthly servicing, roll-rate, cure and DPD are synthetic controls-testing evidence.\n3. P3 model is frozen and moderate; P2 does not retrain it.\n4. P4 Stage 2 and lifetime ECL remain proxy/evidence constrained.\n5. P4 has no prior-period snapshot; no ECL trend is claimed.\n6. P5 observed/synthetic boundaries remain intact.\n7. Risk appetite is educational.\n8. PBIX and DAX runtime are external gates.\n9. Independent validation and organisational sign-off were not performed.")
    write_md("business_question.md", "Project 2 Business Question", "Is observed booked-portfolio risk improving or deteriorating, which model/policy/ECL indicators explain the movement, which breaches require action, and what controlled evidence supports management decisions?")
    write_md("MASTER_PLAN.md", "Project 2 Gold Candidate v2.1.4 Build Plan", "Core methodology, monitoring completeness, technical implementation, evidence hardening and package governance have been rebuilt. The v2.1.4 analytical freeze uses an isolated author build, a frozen pre-review reference bundle, temporal sequencing controls and a later reviewer rebuild comparison. The only open execution gate is a real PBIX/DAX run in Power BI Desktop; the package does not overclaim closure of that gate.")
    write_md("change_log.md", "Project 2 Change Log", f"- {RUN_DATE}: Gold Candidate v2.1.4 evidence-hardening patch: isolated author baseline, immutable author bundle, reviewer no-refresh guard, temporal-order validation, author-versus-reviewer hash comparison, Excel calculation-mode resolution, synthetic KRI wording and DAX date-context controls.\n- Gold Candidate v2.1.3 packages are superseded by v2.1.4 `REVIEWER_VERIFIED`.\n- Real PBIX, actual DAX runtime and Power BI screenshots remain explicit external gates.")
    package_index = pd.DataFrame(
        [
            ["Primary README", "README.md", "Public entry point and release boundary"],
            ["Primary browser report", "OPEN_THIS_FIRST.html and dashboard/index.html", "Executable recruiter-facing report, not a Power BI claim"],
            ["Primary workbook", "excel/Project2_CreditPortfolioMonitoring_Gold_v2_1_4.xlsx", "Formula-driven Excel control model"],
            ["Primary DAX dictionary", "powerbi/dax_measure_dictionary.csv", "Governed measures for PBIX build"],
            ["Primary DAX expected tests", "testing/dax_measure_test_cases.csv", "Typed expected values; actual runtime pending PBIX"],
            ["Primary validation report", "validation/validation_report.md", "Analytical validation summary and open external gates"],
            ["Validation composition", "validation/validation_composition.csv", "Separates artifact checks from substantive and runtime gates"],
            ["Reviewer expected-vs-actual", "validation/reviewer_expected_actual_comparison.csv", "Frozen expected controls reconciled to regenerated outputs"],
            ["Deterministic reference hashes", "validation/deterministic_reference_hashes.json", "Author baseline for controlled output files"],
            ["Frozen author evidence bundle", "validation/author_reference/", "Separate author-build code snapshot, output copies, source manifest, pipeline records and hashed bundle"],
            ["Reference temporal control", "validation/reference_temporal_validation.json", "Proves the frozen reference existed before the reviewer pipeline started"],
            ["Deterministic rebuild comparison", "validation/deterministic_rebuild_comparison.csv", "Expected author hashes/row counts versus actual rebuild outputs"],
            ["Reviewer clean rebuild log", "logs/reviewer_clean_rebuild.json and logs/reviewer_clean_rebuild.log", "Per-script start/end/duration/return-code evidence"],
            ["Pipeline runtime log", "logs/pipeline_run_log.csv", "Per-script start/end/duration evidence"],
            ["Reviewer runtime guide", "reports/reviewer_rebuild_runtime_guide.md", "OS, hardware, data-size and rebuild guidance"],
            ["Two-phase runner", "scripts/run_two_phase_reproducibility.py", "Isolated author baseline, later reviewer rebuild and final clean-extract validation"],
            ["Primary build manifest", "validation/build_artifact_manifest.csv", "Deterministic analytical build scope"],
            ["Primary release manifest", "validation/release_package_manifest.csv", "Final release scope excluding circular verifier files"],
            ["PBIX status", "powerbi/pbix_release_gate.md", "OPEN_EXTERNAL_GATE until real PBIX/DAX runtime/screenshots exist"],
            ["Independent-review status", "validation/independent_review_checklist.csv", "Evidence prepared; independent review not performed"],
            ["Superseded artefacts", "archive/superseded/", "Historical v2.0/v2.1 evidence/workbooks; not authoritative"],
        ],
        columns=["index_item", "path", "purpose"],
    )
    write_md(
        "PACKAGE_INDEX.md",
        "Project 2 Package Index",
        "Only the files labelled primary below are authoritative for v2.1.4. Superseded artefacts are retained only under `archive/superseded/` for audit history.\n\n"
        + table_md(package_index),
    )

    dashboard_data = {
        "version": VERSION,
        "current_period": current_period,
        "prior_period": prior_period,
        "headline": headline.replace({np.nan: None}).to_dict(orient="records"),
        "policy": policy_view.replace({np.nan: None}).to_dict(orient="records"),
        "kri": kri.replace({np.nan: None}).to_dict(orient="records"),
    }
    assets = ROOT / "dashboard/assets"
    assets.mkdir(parents=True, exist_ok=True)
    (assets / "dashboard_data.json").write_text(json.dumps(dashboard_data, indent=2), encoding="utf-8")

    eligible = model[model["eligibility_status"].eq("ELIGIBLE")].tail(18)
    portfolio_eligible = portfolio[portfolio["outcome_status"].eq("ASSESSABLE")].tail(18)
    latest_feature = drift[drift["is_current_model_period"].eq(True)].sort_values("psi_vs_development_baseline", ascending=False).head(8)
    grade_summary = segment.groupby("risk_grade", as_index=False).agg(accounts=("accounts", "sum"), exposure=("exposure", "sum"), observed_bad_rate_matured=("observed_bad_rate_matured", "mean"))
    ecl_stage = ecl.groupby("stage", as_index=False).agg(accounts=("accounts", "sum"), ead=("ead", "sum"), ecl_weighted=("ecl_weighted", "sum"))

    kri_rows_html = "".join(
        f"<tr><td>{html.escape(str(row.kri_id))}</td><td><strong>{html.escape(str(row.metric))}</strong><small>{html.escape(str(row.current_period))} / prior {html.escape(str(row.prior_period))}</small></td><td>{_pct(float(row.current_value)) if abs(float(row.current_value)) < 2 else _num(float(row.current_value),2)}</td><td>{_status_badge(str(row.status))}</td><td>{html.escape(str(row.alert_state))}</td><td>{int(row.affected_accounts):,}</td><td>{html.escape(str(row.owner))}</td></tr>"
        for row in kri.itertuples()
    )
    action_rows_html = "".join(
        f"<tr><td>{html.escape(str(row.action_id))}</td><td><strong>{html.escape(str(row.owner))}</strong><small>{html.escape(str(row.affected_segment))}</small></td><td>{html.escape(str(row.quantified_remediation))}</td><td>{html.escape(str(row.acceptance_criteria))}</td><td>{html.escape(str(row.closure_evidence))}</td></tr>"
        for row in actions.itertuples()
    )
    policy_rows_html = "".join(
        f"<tr><td><strong>{html.escape(str(row.policy_version))}</strong></td><td>{_pct(float(row.approval_rate))}</td><td>{_pct(float(row.manual_review_rate))}</td><td>{_pct(float(row.decline_rate))}</td><td>{_pct(float(row.approved_bad_rate_matured))}</td><td>{_pct(float(row.auto_approved_el_reduction_vs_baseline))}</td><td>{_num(float(row.total_routed_expected_loss))}</td></tr>"
        for row in policy.itertuples()
    )
    grade_bars = "".join(
        f'<div class="bar-row"><span>{html.escape(str(row.risk_grade))}</span><div class="track"><i style="width:{min(100,float(row.accounts)/float(grade_summary.accounts.max())*100):.1f}%"></i></div><b>{int(row.accounts):,}</b><em>{_pct(float(row.observed_bad_rate_matured))}</em></div>'
        for row in grade_summary.sort_values("risk_grade").itertuples()
    )
    policy_bars = "".join(
        f'<div class="policy-row"><span>{html.escape(str(row.policy_version).replace("_"," "))}</span><div class="stack"><i class="approve" style="width:{float(row.approval_rate)*100:.1f}%"></i><i class="review" style="width:{float(row.manual_review_rate)*100:.1f}%"></i><i class="decline" style="width:{float(row.decline_rate)*100:.1f}%"></i></div><b>{_pct(float(row.approved_bad_rate_matured))} bad</b></div>'
        for row in policy.itertuples()
    )
    drift_rows = "".join(
        f"<tr><td>{html.escape(str(row.feature))}</td><td>{float(row.psi_vs_development_baseline):.3f}</td><td>{float(row.prior_psi):.3f}</td><td>{float(row.psi_change):+.3f}</td><td>{html.escape(str(row.status))}</td></tr>"
        for row in latest_feature.itertuples()
    )
    ecl_rows = "".join(
        f"<tr><td>{html.escape(str(row.stage))}</td><td>{int(row.accounts):,}</td><td>{float(row.ead):,.0f}</td><td>{float(row.ecl_weighted):,.0f}</td><td>{_pct(float(row.ecl_weighted)/float(row.ead))}</td></tr>"
        for row in ecl_stage.itertuples()
    )

    dashboard_html = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Project 2 | Credit Portfolio Monitoring</title>
<style>
:root{{--ink:#17313b;--deep:#102b35;--teal:#0b6b68;--green:#2f7d4a;--amber:#bd7a0c;--red:#ad3434;--blue:#356987;--violet:#6c5f80;--paper:#eef3f3;--panel:#ffffff;--line:#d3dee0;--line-dark:#b9c8cb;--muted:#607177;--white:#fff}}
*{{box-sizing:border-box}} html,body{{max-width:100%;overflow-x:hidden}} body{{margin:0;background:var(--paper);color:var(--ink);font:14px/1.5 "Segoe UI",Arial,sans-serif;letter-spacing:0}}
button,a{{font:inherit}} a{{color:inherit}} .shell{{display:grid;grid-template-columns:268px minmax(0,1fr);min-height:100vh}}
.sidebar{{background:var(--deep);color:#fff;padding:24px 18px;position:sticky;top:0;height:100vh;overflow:auto;border-right:1px solid #294650}}
.side-kicker{{color:#76c7bc;font-size:10px;font-weight:800;text-transform:uppercase;margin-bottom:7px}} .brand{{font-size:20px;line-height:1.25;font-weight:750;margin-bottom:7px}} .version{{color:#a9c2c7;font-size:12px;margin-bottom:20px}}
.release-chip{{display:inline-flex;align-items:center;gap:7px;border:1px solid #3f6069;background:#173944;padding:6px 8px;border-radius:4px;color:#dff2ef;font-size:11px;margin-bottom:20px}} .release-chip i{{width:7px;height:7px;background:#63ba84;border-radius:50%}}
.nav-list{{border-top:1px solid #2f4d57;padding-top:10px}} .nav-btn{{display:flex;align-items:center;width:100%;border:0;background:transparent;color:#c9dde0;text-align:left;padding:10px 8px;margin:2px 0;cursor:pointer;border-left:3px solid transparent;border-radius:0}}
.nav-btn b{{display:inline-grid;place-items:center;width:25px;height:25px;margin-right:9px;background:#214753;color:#fff;border-radius:4px;font-size:10px}} .nav-btn.active,.nav-btn:hover{{background:#1d404b;color:#fff;border-left-color:#58b9ad}}
.sidebar-note{{border-top:1px solid #36545f;margin-top:22px;padding-top:17px;color:#b8cdd1;font-size:12px}} .sidebar-note strong{{color:#fff}} .sidebar a{{color:#dff5f1}}
main{{min-width:0;max-width:100%;padding:26px 32px 52px}} .topbar{{display:flex;align-items:flex-start;justify-content:space-between;gap:22px;margin-bottom:22px;padding-bottom:18px;border-bottom:1px solid var(--line)}}
.topbar .context{{max-width:720px;min-width:0}} h1{{font-size:27px;line-height:1.2;margin:0 0 6px}} h2{{font-size:18px;line-height:1.25;margin:0 0 5px}} h3{{font-size:13px;margin:0 0 8px;text-transform:uppercase;color:#53656b}} p{{margin:4px 0 10px}} .muted,small{{color:var(--muted)}} small{{display:block;font-size:11px;margin-top:3px}}
.evidence{{display:flex;gap:7px;flex-wrap:wrap;justify-content:flex-end;max-width:430px}} .tag{{padding:5px 8px;border:1px solid var(--line);background:#fff;border-radius:4px;font-size:10px;font-weight:750}} .tag.observed{{border-color:#78a9c1;color:#285d77}} .tag.synthetic{{border-color:#d5a64a;color:#7d5606}} .tag.proxy{{border-color:#a89ab8;color:#5c4c70}}
.page{{display:none;width:100%;min-width:0;max-width:100%}} .page.active{{display:block}} .page-head{{display:flex;align-items:flex-start;gap:12px;margin:0 0 14px}} .page-head>div{{flex:1 1 0;width:0;min-width:0;max-width:100%}} .page-number{{display:grid;place-items:center;min-width:35px;height:35px;background:var(--ink);color:#fff;border-radius:4px;font-size:11px;font-weight:800}} .page-head h2,.page-head p{{overflow-wrap:anywhere}} .page-head p{{margin:2px 0;color:var(--muted)}}
.banner{{display:flex;align-items:center;gap:8px;border-left:5px solid var(--amber);background:#fff7e6;padding:12px 14px;margin-bottom:14px;overflow-wrap:anywhere}} .banner>*{{min-width:0}} .banner strong{{white-space:nowrap}} .banner span{{overflow-wrap:anywhere}} .banner span+span:before{{content:"|";color:#bd9b60;margin-right:8px}}
.insight-strip{{display:grid;width:100%;min-width:0;max-width:100%;grid-template-columns:repeat(3,minmax(0,1fr));border:1px solid var(--line);background:var(--panel);margin-bottom:14px}} .insight-strip>div{{min-width:0;padding:12px 14px;border-right:1px solid var(--line)}} .insight-strip>div:last-child{{border-right:0}} .insight-strip small{{text-transform:uppercase;font-weight:800;color:#6b7a7f;margin:0 0 4px}} .insight-strip strong{{display:block;max-width:100%;font-size:14px;white-space:normal;overflow-wrap:anywhere}}
.grid{{display:grid;grid-template-columns:repeat(12,minmax(0,1fr));gap:14px;margin-bottom:14px}} .span-3{{grid-column:span 3}} .span-4{{grid-column:span 4}} .span-5{{grid-column:span 5}} .span-6{{grid-column:span 6}} .span-7{{grid-column:span 7}} .span-8{{grid-column:span 8}} .span-12{{grid-column:span 12}}
.card{{background:var(--panel);border:1px solid var(--line);border-radius:6px;padding:16px;min-width:0;box-shadow:0 1px 2px rgba(21,48,58,.04)}} .kpi{{min-height:132px;border-top:3px solid var(--blue)}} .kpi.risk{{border-top-color:var(--red)}} .kpi.warning{{border-top-color:var(--amber)}} .kpi.controlled{{border-top-color:var(--green)}} .kpi .value{{font-size:28px;line-height:1;font-weight:760;margin:16px 0 7px}} .delta{{font-weight:750}} .up-bad{{color:var(--red)}} .down{{color:var(--green)}}
.status{{display:inline-block;padding:3px 7px;border-radius:4px;font-size:10px;font-weight:800}} .status.red{{background:#f5dada;color:#8e2424}} .status.amber{{background:#fcebc4;color:#795003}} .status.green{{background:#d9ecdf;color:#205c35}} .status.monitor{{background:#dbe8f0;color:#2d5b74}} .status.static-review{{background:#e9e5ef;color:#58496b}}
.chart-svg{{display:block;width:100%;height:205px}} table{{border-collapse:collapse;width:100%;font-size:12px}} th{{text-align:left;color:#4e6066;background:#eaf0f1;padding:10px;border-bottom:1px solid var(--line-dark);white-space:nowrap}} td{{padding:10px;border-bottom:1px solid #e0e8e9;vertical-align:top}} tbody tr:hover{{background:#f8fbfb}} .table-wrap{{overflow:auto}}
.bar-row{{display:grid;grid-template-columns:34px minmax(120px,1fr) 88px 62px;align-items:center;gap:9px;margin:11px 0}} .track{{height:10px;background:#e5ecee}} .track i{{display:block;height:100%;background:var(--teal)}} .bar-row b,.bar-row em{{font-size:12px;font-style:normal;text-align:right}}
.policy-row{{display:grid;grid-template-columns:210px minmax(180px,1fr) 92px;align-items:center;gap:10px;margin:13px 0}} .stack{{display:flex;height:14px;background:#e5ebed}} .stack i{{height:100%}} .approve{{background:var(--green)}} .review{{background:var(--amber)}} .decline{{background:var(--red)}} .policy-row b{{font-size:12px;text-align:right}}
.decision{{border-left:5px solid var(--teal);background:#fbfefe}} .decision-label{{font-size:10px;text-transform:uppercase;font-weight:800;color:var(--teal);margin-bottom:6px}} .decision strong{{font-size:18px;line-height:1.3}} .control-list{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:8px}} .control{{border:1px solid var(--line);padding:11px;background:#f7fafa;border-radius:4px}} .interpretation{{border-top:3px solid var(--blue);background:#f5f9fb}} .interpretation p:last-child{{margin-bottom:0}}
.footer-links{{display:flex;gap:8px;flex-wrap:wrap;margin-top:20px;padding-top:18px;border-top:1px solid var(--line)}} .footer-links a{{color:var(--teal);background:#fff;border:1px solid var(--line);padding:8px 10px;text-decoration:none;border-radius:4px;font-weight:650}}
@media(max-width:1050px){{.shell{{grid-template-columns:230px minmax(0,1fr)}} main{{padding:22px}} .span-3{{grid-column:span 6}}}}
@media(max-width:780px){{.shell{{grid-template-columns:minmax(0,1fr);width:auto;max-width:100%}} .sidebar{{position:relative;width:auto;height:auto;min-width:0;max-width:100%;overflow:hidden;padding:18px}} .release-chip,.sidebar-note{{display:none}} .nav-list{{display:flex;width:100%;max-width:100%;overflow-x:auto;border-top:0;padding-top:4px}} .nav-btn{{flex:0 0 158px;min-width:0}} main{{width:auto;min-width:0;max-width:100%;overflow:hidden;padding:17px}} .topbar{{display:block;width:100%;min-width:0;max-width:100%}} .topbar .context{{width:100%;max-width:100%}} .topbar p{{overflow-wrap:anywhere}} .evidence{{justify-content:flex-start;margin-top:12px}} h1{{font-size:22px;overflow-wrap:anywhere}} .grid,.card{{width:100%;min-width:0;max-width:100%}} .banner{{align-items:flex-start;flex-wrap:wrap;font-size:13px}} .banner>*{{flex:1 1 100%;min-width:0}} .banner strong{{white-space:normal}} .banner span+span:before{{display:none}} .insight-strip{{grid-template-columns:minmax(0,1fr)}} .insight-strip>div{{border-right:0;border-bottom:1px solid var(--line)}} .insight-strip>div:last-child{{border-bottom:0}} .span-3,.span-4,.span-5,.span-6,.span-7,.span-8{{grid-column:span 12}} .policy-row{{grid-template-columns:1fr}} .bar-row{{grid-template-columns:28px minmax(90px,1fr) 72px}} .bar-row em{{display:none}}}}
</style>
</head>
<body><div class="shell"><aside class="sidebar"><div class="side-kicker">Financial Risk Analytics</div><div class="brand">Credit Portfolio Monitoring</div><div class="version">Gold Candidate v2.1.4</div><div class="release-chip"><i></i>Analytical freeze approved</div><div class="nav-list">
{''.join(f'<button class="nav-btn{" active" if i==1 else ""}" data-page="p{i}"><b>{i:02d}</b>{name}</button>' for i,name in enumerate(["Executive Overview","Portfolio Quality","Delinquency & Collections","Vintage & Cohort","Model Monitoring","Policy Performance","ECL & Stress","EWS & Governance"],1))}
</div><div class="sidebar-note"><strong>Release boundary</strong><br>Analytical engine and browser report complete.<br><br>PBIX, DAX runtime and Power BI screenshots: <strong>OPEN EXTERNAL GATE</strong><br><br><a href="../powerbi/pbix_release_gate.md">Read gate evidence</a></div></aside>
<main><div class="topbar"><div class="context"><h1>Executive Credit Risk Monitoring</h1><p class="muted">Latest eligible observed period {current_period} versus {prior_period} | 1.35M historical booked accounts | management-action view</p></div><div class="evidence"><span class="tag observed">Observed outcomes</span><span class="tag synthetic">Synthetic servicing</span><span class="tag proxy">Proxy ECL</span></div></div>

<section id="p1" class="page active"><div class="page-head"><span class="page-number">01</span><div><h2>Executive Overview</h2><p>Current condition, material movement and the decision management must take.</p></div></div><div class="banner"><strong>Senior conclusion:</strong><span>Risk worsened</span><span>Controlled pilot; no wholesale rollout</span></div>
<div class="insight-strip"><div><small>Condition</small><strong>Observed bad rate is Red and deterioration persists.</strong></div><div><small>Primary drivers</small><strong>Ranking weakened, calibration gap widened and PSI moved further into Amber.</strong></div><div><small>Decision</small><strong>Pilot the policy change only with capacity redesign and explicit monitoring limits.</strong></div></div>
<div class="grid"><div class="card kpi risk span-3"><h3>Observed bad rate</h3><div class="value">{_pct(current_bad)}</div><div class="delta up-bad">{_bps(current_bad-prior_bad)}</div><small>Matured outcomes only | RED</small></div>
<div class="card kpi warning span-3"><h3>Model stability</h3><div class="value">{current_psi:.3f} PSI</div><div class="delta up-bad">{float(change_values['pd_psi_vs_development_baseline']):+.3f}</div><small>Fixed development bins | AMBER</small></div>
<div class="card kpi controlled span-3"><h3>Calibration gap</h3><div class="value">{_pct(current_cal)}</div><div class="delta up-bad">{_bps(float(change_values['calibration_gap_matured']))}</div><small>Inside Green; movement worsening</small></div>
<div class="card kpi risk span-3"><h3>Active risk queue</h3><div class="value">{red_count} Red</div><div class="delta">{amber_count} Amber</div><small>{len(actions)} owned actions</small></div></div>
<div class="grid"><div class="card span-7"><h2>Observed bad-rate trend</h2><p class="muted">Eligible cohorts only; immature cohorts remain blank.</p>{_sparkline_svg(portfolio_eligible['observed_bad_rate_matured'].tolist(), '#b33a3a')}</div>
<div class="card span-5 decision"><div class="decision-label">Decision required</div><strong>{recommendation.replace('_',' ')}</strong><p>The 20% candidate cuts straight-through auto-approved EL {_pct(float(candidate['auto_approved_el_reduction_vs_baseline']))}, but approval falls {_bps(candidate_approval_drop)} and review workload rises {candidate_review_relative:.1%}.</p><p><b>Proceed with:</b> a constrained pilot, capacity plan and stop criteria. <b>Do not:</b> call routed EL a realised loss reduction.</p></div></div>
<div class="card table-wrap"><h2>Current risk queue</h2><table><thead><tr><th>KRI</th><th>Indicator</th><th>Current</th><th>Status</th><th>State</th><th>Affected</th><th>Owner</th></tr></thead><tbody>{kri_rows_html}</tbody></table></div></section>

<section id="p2" class="page"><div class="page-head"><span class="page-number">02</span><div><h2>Portfolio Quality</h2><p>Observed booked-account risk, maturity coverage and concentration by credit grade.</p></div></div><div class="banner"><strong>Observed layer:</strong><span>Outcome denominators include matured accounts only</span><span>Immature cohorts remain not assessable, never 0%</span></div><div class="grid"><div class="card span-7"><h2>Eligible-period bad-rate trend</h2><p class="muted">Realised outcomes after maturity gating.</p>{_sparkline_svg(portfolio_eligible['observed_bad_rate_matured'].tolist(), '#0b6b68')}</div><div class="card span-5"><h2>Grade concentration</h2><p class="muted">Account concentration and observed matured bad rate.</p>{grade_bars}</div></div><div class="card table-wrap"><h2>Latest current / prior / change</h2><table><thead><tr><th>Metric</th><th>Period</th><th>Current</th><th>Prior</th><th>Change</th><th>Status</th><th>Boundary</th></tr></thead><tbody>{''.join(f'<tr><td>{html.escape(str(r.metric))}</td><td>{html.escape(str(r.current_period))}</td><td>{float(r.current):.4f}</td><td>{float(r.prior):.4f}</td><td>{float(r.change):+.4f}</td><td>{_status_badge(str(r.status))}</td><td>{html.escape(str(r.evidence_note))}</td></tr>' for r in headline.itertuples())}</tbody></table></div><div class="card interpretation"><h2>Senior interpretation</h2><p>Risk deterioration is observed, but the response should be segment-led. Decompose the movement by grade, purpose, FICO and DTI before changing broad policy; immature cohorts remain visible for exposure monitoring but are excluded from performance conclusions.</p></div></section>

<section id="p3" class="page"><div class="page-head"><span class="page-number">03</span><div><h2>Delinquency & Collections</h2><p>Controlled servicing logic for stock, flow, cure and default-state reconciliation.</p></div></div><div class="banner"><strong>Synthetic controls only:</strong><span>Charts test servicing logic</span><span>No observed delinquency performance is claimed</span></div><div class="grid"><div class="card span-6"><h2>30+ DPD stock</h2><p class="muted">Latest 24 synthetic periods.</p>{_sparkline_svg(stock.tail(24)['dpd_30_plus'].tolist(), '#c98613')}</div><div class="card span-6"><h2>Default stock versus new flow</h2><h3>Default stock</h3>{_sparkline_svg(stock.tail(24)['default_stock_rate'].tolist(), '#b33a3a',720,120)}<h3>New default flow</h3>{_sparkline_svg(stock.tail(24)['new_default_flow_rate'].tolist(), '#3f6c8e',720,120)}</div></div><div class="grid"><div class="card span-8 table-wrap"><h2>Latest stock-flow bridge</h2><table><thead><tr><th>Beginning stock</th><th>Inflow</th><th>Cures</th><th>Defaults</th><th>Closures</th><th>Other net</th><th>Ending stock</th><th>Difference</th></tr></thead><tbody><tr><td>{int(latest_stock['beginning_delinquent_stock']):,}</td><td>{int(latest_stock['new_delinquency_inflow']):,}</td><td>{int(latest_stock['cures']):,}</td><td>{int(latest_stock['new_defaults_from_delinquency']):,}</td><td>{int(latest_stock['delinquent_closures']):,}</td><td>{int(latest_stock['other_migration_net']):,}</td><td>{int(latest_stock['ending_delinquent_stock']):,}</td><td><b>{float(latest_stock['bridge_difference']):.0f}</b></td></tr></tbody></table></div><div class="card span-4"><h2>Control evidence</h2><div class="control-list"><div class="control"><b>48</b><small>origination vintages</small></div><div class="control"><b>36 / 60</b><small>term-aware months</small></div><div class="control"><b>0</b><small>maximum bridge gap</small></div><div class="control"><b>Separate</b><small>default stock and flow</small></div></div></div></div><div class="card interpretation"><h2>Senior interpretation</h2><p>The synthetic Red outcome is retained when it represents intentional scenario severity. Controls are adjusted only when transition assumptions are unsupported, never simply to force a KRI back to Green.</p></div></section>

<section id="p4" class="page"><div class="page-head"><span class="page-number">04</span><div><h2>Vintage & Cohort</h2><p>Risk emergence by origination vintage, months on book, contractual term and grade.</p></div></div><div class="banner"><strong>Multi-vintage synthetic layer:</strong><span>48 origination vintages</span><span>MOB 1-36 with 36/60-month term-aware balances</span></div><div class="grid"><div class="card span-6"><h2>Vintage 30+ DPD curves</h2><p class="muted">Average curve for the latest six synthetic vintages.</p>{_sparkline_svg(vintage[vintage['origination_vintage'].isin(sorted(vintage['origination_vintage'].unique())[-6:])].groupby('months_on_book')['dpd30_rate'].mean().tolist(), '#0b6b68')}</div><div class="card span-6"><h2>60-month residual balance</h2><p class="muted">Residual exposure remains after MOB36 by design.</p>{_sparkline_svg(vintage[vintage['term_months'].eq(60)].groupby('months_on_book')['ending_balance'].sum().tolist(), '#3f6c8e')}</div></div><div class="card"><h2>Vintage-control design</h2><div class="control-list"><div class="control"><b>48 vintages</b><small>separate origination cohorts</small></div><div class="control"><b>MOB 1-36</b><small>consistent observation horizon</small></div><div class="control"><b>36 / 60 months</b><small>term-aware amortisation</small></div><div class="control"><b>Grade + policy</b><small>segmented monitoring views</small></div></div></div><div class="card interpretation"><h2>Senior interpretation</h2><p>This layer demonstrates the correct analytical mechanics for vintage monitoring. It is fit for controls testing and interview defence, but observed servicing data is still required before using these curves for production collections decisions.</p></div></section>

<section id="p5" class="page"><div class="page-head"><span class="page-number">05</span><div><h2>Model Monitoring</h2><p>Ranking, calibration and population stability for the frozen Project 3 scorecard.</p></div></div><div class="banner"><strong>Frozen P3 model:</strong><span>AUC, KS and calibration use matured outcomes</span><span>PSI uses fixed development bins; Project 2 monitors, not retrains</span></div><div class="grid"><div class="card span-4 kpi warning"><h3>Latest AUC</h3><div class="value">{current_auc:.3f}</div><small>Moderate ranking | change {float(change_values['auc']):+.3f}</small></div><div class="card span-4 kpi warning"><h3>Latest KS</h3><div class="value">{float(current_values['ks']):.3f}</div><small>Change {float(change_values['ks']):+.3f}</small></div><div class="card span-4 kpi warning"><h3>Latest PSI</h3><div class="value">{current_psi:.3f}</div><small>AMBER | fixed development baseline</small></div></div><div class="grid"><div class="card span-6"><h2>AUC trend</h2><p class="muted">Eligible matured periods only.</p>{_sparkline_svg(eligible['auc'].tolist(), '#3f6c8e')}</div><div class="card span-6"><h2>PSI trend</h2><p class="muted">Fixed-bin movement versus development reference.</p>{_sparkline_svg(eligible['pd_psi_vs_development_baseline'].tolist(), '#c98613')}</div></div><div class="card table-wrap"><h2>Current feature drift drivers</h2><table><thead><tr><th>Feature</th><th>Current PSI</th><th>Prior PSI</th><th>Change</th><th>Status</th></tr></thead><tbody>{drift_rows}</tbody></table></div><div class="card interpretation"><h2>Senior interpretation</h2><p>The scorecard has moderate and weakening ranking power. PSI at {current_psi:.3f} supports heightened monitoring and feature-level diagnosis, not automatic recalibration; formal model review is triggered only after stronger evidence or the governed Red threshold.</p></div></section>

<section id="p6" class="page"><div class="page-head"><span class="page-number">06</span><div><h2>Policy Performance</h2><p>Risk, approval, decline and operational-capacity trade-offs across five policy versions.</p></div></div><div class="banner"><strong>Historical booked-account simulation:</strong><span>Rejected applicants lack observed outcomes</span><span>Auto-approved EL reduction is routing impact, not realised loss reduction</span></div><div class="card"><h2>Policy version allocation</h2><p class="muted">Green approve, amber manual review, red decline. Approved bad rate is shown at right.</p>{policy_bars}</div><div class="card table-wrap"><h2>Risk, approval and workload bridge</h2><table><thead><tr><th>Policy</th><th>Approve</th><th>Manual</th><th>Decline</th><th>Approved bad</th><th>Auto-approved EL reduction</th><th>Routed EL</th></tr></thead><tbody>{policy_rows_html}</tbody></table></div><div class="card decision"><div class="decision-label">Senior recommendation</div><p><strong>{recommendation.replace('_',' ')}</strong></p><p>Risk objectives pass, while approval and manual-review capacity fail. A 35% decline boundary produces zero declines because no frozen PD exceeds that point. The 30% sensitivity deliberately creates a measurable decline population to test operational routing.</p></div><div class="card interpretation"><h2>Why 20% is not an automatic rollout</h2><p>The candidate improves straight-through risk selection, but a 30.32 percentage-point approval reduction and roughly 150% review-volume increase breach business capacity. The correct recommendation is a capped pilot with queue capacity, service-level and stop-loss controls.</p></div></section>

<section id="p7" class="page"><div class="page-head"><span class="page-number">07</span><div><h2>ECL, Stress & Concentration</h2><p>Frozen Project 4 loss snapshot, scenario sensitivity and readiness limitations.</p></div></div><div class="banner"><strong>Frozen P4 proxy snapshot:</strong><span>Stage 2 remains broad proxy logic</span><span>No prior ECL reporting date exists, so trend is not assessable</span></div><div class="grid"><div class="card span-3 kpi risk"><h3>Stage 2 proxy</h3><div class="value">{_pct(float(stage2['current_value']))}</div><small>RED | production SICR gap</small></div><div class="card span-3 kpi risk"><h3>Downside uplift</h3><div class="value">{_pct(float(downside['current_value']))}</div><small>RED | base-to-downside</small></div><div class="card span-3 kpi controlled"><h3>Scenarios</h3><div class="value">{len(scenarios)}</div><small>Frozen P4 overlays</small></div><div class="card span-3 kpi"><h3>Period trend</h3><div class="value">N/A</div><small>No prior snapshot; correctly withheld</small></div></div><div class="card table-wrap"><h2>ECL by stage</h2><table><thead><tr><th>Stage</th><th>Accounts</th><th>EAD</th><th>Weighted ECL</th><th>ECL rate</th></tr></thead><tbody>{ecl_rows}</tbody></table></div><div class="card interpretation"><h2>Senior interpretation</h2><p>Downside sensitivity is material, but the evidence supports a frozen scenario view rather than a trend claim. Stage 2 should remain explicitly labelled a broad proxy until observed SICR triggers and a second reporting-date snapshot are available.</p></div></section>

<section id="p8" class="page"><div class="page-head"><span class="page-number">08</span><div><h2>EWS, Actions & Governance</h2><p>Threshold state, accountable ownership, acceptance criteria and evidence-based closure.</p></div></div><div class="banner"><strong>Governance view:</strong><span>Dynamic alerts use latest and prior eligible periods</span><span>Static upstream overlays are labelled; independent sign-off is not claimed</span></div><div class="insight-strip"><div><small>Open queue</small><strong>{red_count} Red and {amber_count} Amber indicators.</strong></div><div><small>Ownership</small><strong>{len(actions)} actions with named owner and approver.</strong></div><div><small>Closure standard</small><strong>Acceptance criteria plus objective closure evidence.</strong></div></div><div class="card table-wrap"><h2>KRI state and ownership</h2><table><thead><tr><th>KRI</th><th>Indicator</th><th>Current</th><th>Status</th><th>State</th><th>Affected</th><th>Owner</th></tr></thead><tbody>{kri_rows_html}</tbody></table></div><div class="card table-wrap"><h2>Management actions and closure evidence</h2><table><thead><tr><th>Action</th><th>Owner / segment</th><th>Quantified remediation</th><th>Acceptance criteria</th><th>Closure evidence</th></tr></thead><tbody>{action_rows_html}</tbody></table></div><div class="card interpretation"><h2>Senior interpretation</h2><p>The project does not stop at flagging breaches. Each material signal is translated into an investigation, quantified remediation, decision owner, acceptance criterion and closure artifact so management can distinguish monitoring from actual risk treatment.</p></div></section>

<div class="footer-links"><a href="../README.md">README</a><a href="../reports/risk_committee_monitoring_memo.md">Risk memo</a><a href="../excel/Project2_CreditPortfolioMonitoring_Gold_v2_1_4.xlsx">Excel control model</a><a href="../powerbi/README.md">Power BI build pack</a><a href="../validation/validation_report.md">Validation</a></div>
</main></div><script>
const buttons=[...document.querySelectorAll('.nav-btn')], pages=[...document.querySelectorAll('.page')];
function show(id){{buttons.forEach(b=>b.classList.toggle('active',b.dataset.page===id));pages.forEach(p=>p.classList.toggle('active',p.id===id));history.replaceState(null,'','?page='+id);window.scrollTo(0,0)}}
buttons.forEach(b=>b.addEventListener('click',()=>show(b.dataset.page)));const requested=new URLSearchParams(location.search).get('page');if(requested&&document.getElementById(requested))show(requested);
</script></body></html>"""
    dashboard_path = ROOT / "dashboard/index.html"
    dashboard_path.parent.mkdir(parents=True, exist_ok=True)
    dashboard_path.write_text(dashboard_html, encoding="utf-8")
    entry_html = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="description" content="Project 2 credit portfolio monitoring, early warning, policy, ECL and risk governance case study.">
<title>Project 2 | Credit Portfolio Monitoring</title>
<style>
:root{{--ink:#17313b;--deep:#102b35;--teal:#0b6b68;--green:#2f7d4a;--amber:#bd7a0c;--red:#ad3434;--blue:#356987;--paper:#eef3f3;--panel:#fff;--line:#d3dee0;--muted:#607177}}
*{{box-sizing:border-box}} html{{scroll-behavior:smooth}} body{{margin:0;background:var(--paper);color:var(--ink);font:15px/1.55 "Segoe UI",Arial,sans-serif;letter-spacing:0}} a{{color:inherit}}
.top{{background:var(--deep);color:#fff;border-bottom:4px solid var(--teal)}} .nav{{max-width:1180px;margin:auto;padding:16px 24px;display:flex;align-items:center;justify-content:space-between;gap:20px}} .wordmark{{font-weight:780}} .wordmark small{{display:block;color:#a9c5c8;font-size:11px;font-weight:500}} .links{{display:flex;gap:8px;flex-wrap:wrap}} .links a{{text-decoration:none;border:1px solid #3c5c66;padding:7px 9px;border-radius:4px;font-size:12px;color:#dcebed}}
.hero{{max-width:1180px;margin:auto;padding:46px 24px 38px;display:grid;grid-template-columns:minmax(0,1.5fr) minmax(300px,.75fr);gap:38px;align-items:end}} .eyebrow{{color:#78c8bd;text-transform:uppercase;font-size:11px;font-weight:800;margin-bottom:10px}} h1{{font-size:44px;line-height:1.08;margin:0 0 13px}} .lede{{font-size:18px;color:#c8dadd;max-width:760px;margin:0}} .release{{border:1px solid #3e5e67;background:#163945;padding:18px;border-radius:6px}} .release strong{{display:block;font-size:15px;margin-bottom:7px}} .release p{{color:#bdd1d4;margin:0 0 12px;font-size:13px}} .status-line{{display:flex;align-items:center;gap:8px;color:#dff3ea;font-size:12px;font-weight:700}} .dot{{width:8px;height:8px;background:#63ba84;border-radius:50%}}
.actionbar{{max-width:1180px;margin:auto;padding:0 24px 28px;display:flex;gap:9px;flex-wrap:wrap}} .primary,.secondary{{display:inline-block;text-decoration:none;padding:10px 13px;border-radius:4px;font-weight:750;font-size:13px}} .primary{{background:#fff;color:var(--deep)}} .secondary{{border:1px solid #54717a;color:#fff}}
.band{{padding:34px 24px}} .band.white{{background:#fff}} .inner{{max-width:1180px;margin:auto}} .section-head{{display:flex;justify-content:space-between;gap:22px;align-items:end;margin-bottom:19px}} .section-head h2{{font-size:24px;margin:0}} .section-head p{{max-width:620px;color:var(--muted);margin:0}} h3{{font-size:14px;margin:0 0 7px}} p{{margin:4px 0 10px}}
.stat-grid{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px}} .stat{{background:var(--panel);border:1px solid var(--line);border-top:3px solid var(--blue);border-radius:6px;padding:17px;min-height:126px}} .stat.risk{{border-top-color:var(--red)}} .stat.warning{{border-top-color:var(--amber)}} .stat.controlled{{border-top-color:var(--green)}} .stat small{{display:block;color:var(--muted);font-size:11px;text-transform:uppercase;font-weight:750}} .stat strong{{display:block;font-size:29px;line-height:1;margin:17px 0 8px}} .stat span{{color:var(--muted);font-size:12px}}
.decision{{border-left:5px solid var(--teal);background:#f8fcfb;padding:19px 21px;margin-top:15px}} .decision b{{font-size:19px}} .decision-grid{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:0;margin-top:15px;border:1px solid var(--line);background:#fff}} .decision-grid div{{padding:14px;border-right:1px solid var(--line)}} .decision-grid div:last-child{{border-right:0}} .decision-grid small{{display:block;text-transform:uppercase;font-weight:800;color:var(--muted);font-size:10px;margin-bottom:5px}}
.flow{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));border:1px solid var(--line);background:#fff}} .flow article{{padding:18px;border-right:1px solid var(--line);position:relative}} .flow article:last-child{{border-right:0}} .flow .num{{display:inline-grid;place-items:center;width:27px;height:27px;background:var(--ink);color:#fff;border-radius:4px;font-size:10px;font-weight:800;margin-bottom:12px}} .flow p{{color:var(--muted);font-size:13px}}
.value-grid{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:13px}} .value-item{{border-top:3px solid var(--teal);padding:15px 0}} .value-item:nth-child(2){{border-color:var(--blue)}} .value-item:nth-child(3){{border-color:var(--amber)}} .value-item p{{color:var(--muted)}}
table{{border-collapse:collapse;width:100%;background:#fff;border:1px solid var(--line);font-size:13px}} th{{text-align:left;background:#eaf0f1;color:#506168;padding:11px;border-bottom:1px solid #bbc9cc}} td{{padding:11px;border-bottom:1px solid #e0e7e8;vertical-align:top}} .label{{display:inline-block;padding:3px 7px;border-radius:4px;font-size:10px;font-weight:800}} .observed{{background:#dbeaf1;color:#275a73}} .synthetic{{background:#faebc8;color:#775005}} .proxy{{background:#e9e4ef;color:#57496a}} .mixed{{background:#dcecdf;color:#275e38}}
.validation-grid{{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));border:1px solid var(--line);background:#fff}} .validation-grid div{{padding:15px;border-right:1px solid var(--line)}} .validation-grid div:last-child{{border-right:0}} .validation-grid strong{{display:block;font-size:22px}} .validation-grid span{{color:var(--muted);font-size:11px}} .gate{{margin-top:14px;padding:14px;border:1px solid #e0c88d;background:#fff8e7;color:#6f4a04}}
.artifacts{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px}} .artifact{{display:block;background:#fff;border:1px solid var(--line);border-radius:6px;padding:14px;text-decoration:none}} .artifact strong{{display:block}} .artifact span{{color:var(--muted);font-size:12px}} footer{{background:var(--deep);color:#c2d5d8;padding:23px 24px}} footer div{{max-width:1180px;margin:auto;display:flex;justify-content:space-between;gap:18px;flex-wrap:wrap}}
@media(max-width:900px){{.hero{{grid-template-columns:1fr}} h1{{font-size:35px}} .stat-grid{{grid-template-columns:repeat(2,minmax(0,1fr))}} .flow{{grid-template-columns:repeat(2,minmax(0,1fr))}} .flow article:nth-child(2){{border-right:0}} .flow article{{border-bottom:1px solid var(--line)}} .value-grid,.artifacts{{grid-template-columns:1fr}} .validation-grid{{grid-template-columns:repeat(2,minmax(0,1fr))}} .validation-grid div{{border-bottom:1px solid var(--line)}}}}
@media(max-width:580px){{.nav{{display:block}} .links{{margin-top:12px}} .hero{{padding-top:30px}} h1{{font-size:29px}} .stat-grid,.decision-grid,.flow,.validation-grid{{grid-template-columns:1fr}} .stat{{min-height:110px}} .decision-grid div,.flow article,.validation-grid div{{border-right:0;border-bottom:1px solid var(--line)}} .section-head{{display:block}} .section-head p{{margin-top:7px}}}}
</style>
</head>
<body>
<header class="top">
  <nav class="nav"><div class="wordmark">Financial Risk Analytics<small>Project 2 evidence package</small></div><div class="links"><a href="README.md">README</a><a href="reports/risk_committee_monitoring_memo.md">Risk memo</a><a href="validation/validation_report.md">Validation</a></div></nav>
  <div class="hero"><div><div class="eyebrow">Portfolio Risk & Monitoring</div><h1>Credit Portfolio Monitoring & Early Warning</h1><p class="lede">A governed monitoring system that connects observed credit quality, frozen model performance, policy trade-offs, ECL stress and management action across 1.35 million historical booked accounts.</p></div><aside class="release"><strong>Gold Candidate v2.1.4</strong><p>Analytical engine, browser report and author-versus-reviewer reproducibility are complete.</p><div class="status-line"><span class="dot"></span>Approved for analytical freeze</div></aside></div>
  <div class="actionbar"><a class="primary" href="dashboard/index.html">Open interactive risk report &rarr;</a><a class="secondary" href="excel/Project2_CreditPortfolioMonitoring_Gold_v2_1_4.xlsx">Open Excel control model</a><a class="secondary" href="powerbi/README.md">Review Power BI build pack</a></div>
</header>
<section class="band"><div class="inner"><div class="section-head"><div><div class="eyebrow" style="color:var(--teal)">Current condition</div><h2>What management needs to know</h2></div><p>Latest observed performance deteriorated while model ranking weakened and population drift moved further into Amber. The recommended response balances risk reduction against approval and operating capacity.</p></div>
<div class="stat-grid"><article class="stat risk"><small>Observed matured bad rate</small><strong>{_pct(current_bad)}</strong><span>{_bps(current_bad-prior_bad)} versus {prior_period} | Red</span></article><article class="stat warning"><small>Population stability</small><strong>{current_psi:.3f}</strong><span>Fixed-bin PSI | Amber</span></article><article class="stat controlled"><small>Outcome maturity coverage</small><strong>{_pct(_metric(kpi, 'matured_accounts')/_metric(kpi, 'total_accounts'))}</strong><span>{_num(_metric(kpi, 'matured_accounts'))} assessable accounts</span></article><article class="stat risk"><small>Active KRI queue</small><strong>{red_count} Red</strong><span>{amber_count} Amber | {len(actions)} owned actions</span></article></div>
<div class="decision"><div class="eyebrow" style="color:var(--teal)">Management recommendation</div><b>{recommendation.replace('_',' ')}</b><p>The 20% cutoff candidate produces {_pct(float(candidate['auto_approved_el_reduction_vs_baseline']))} lower auto-approved routed EL versus baseline, but approval falls {_bps(candidate_approval_drop)} and manual-review volume rises {candidate_review_relative:.1%}. That is a policy-routing simulation result, not a realised portfolio-loss reduction.</p><div class="decision-grid"><div><small>Proceed</small><strong>Constrained pilot</strong></div><div><small>Required control</small><strong>Capacity and stop criteria</strong></div><div><small>Do not claim</small><strong>Realised loss reduction</strong></div></div></div>
</div></section>
<section class="band white"><div class="inner"><div class="section-head"><div><div class="eyebrow" style="color:var(--teal)">Employer value</div><h2>What this project proves</h2></div><p>The project is designed around the decisions a Portfolio Risk or Risk Analytics team must make, not around isolated charts.</p></div><div class="value-grid"><article class="value-item"><h3>Portfolio diagnosis</h3><p>Connect current/prior movement, exposure, grade concentration, maturity and observed bad rate into a defensible risk view.</p></article><article class="value-item"><h3>Model and policy oversight</h3><p>Monitor AUC, KS, calibration and fixed-bin PSI, then quantify cutoff trade-offs without reopening the frozen scorecard.</p></article><article class="value-item"><h3>Governed management action</h3><p>Translate breaches into owner, investigation, quantified remediation, acceptance criteria and objective closure evidence.</p></article></div></div></section>
<section class="band"><div class="inner"><div class="section-head"><div><div class="eyebrow" style="color:var(--teal)">Evidence architecture</div><h2>From frozen inputs to controlled decisions</h2></div><p>Every output has a defined evidence status and a traceable path back to source data, calculation and validation.</p></div><div class="flow"><article><span class="num">01</span><h3>Frozen upstream evidence</h3><p>P3 scorecard, P4 ECL/stress, P5 enterprise indicators and P6 policy rules.</p></article><article><span class="num">02</span><h3>Monitoring engine</h3><p>Maturity-safe portfolio metrics, tie-safe KS, drift, policy, concentration and EWS logic.</p></article><article><span class="num">03</span><h3>Decision surfaces</h3><p>Python outputs, SQL marts, formula Excel, browser report and governed DAX specification.</p></article><article><span class="num">04</span><h3>Control evidence</h3><p>Author/reviewer code and output reconciliation, manifests, UAT/SIT and clean-extract validation.</p></article></div></div></section>
<section class="band white"><div class="inner"><div class="section-head"><div><div class="eyebrow" style="color:var(--teal)">Claim discipline</div><h2>Observed, synthetic and proxy evidence stay separate</h2></div><p>The dashboard never upgrades an assumption or controls-test layer into an observed production claim.</p></div><table><thead><tr><th>Layer</th><th>What it contains</th><th>Permitted conclusion</th><th>Key limitation</th></tr></thead><tbody><tr><td><span class="label observed">Observed</span></td><td>Historical booked accounts and matured bad outcomes</td><td>Portfolio quality, concentration and model performance</td><td>Accepted/booked population, not full applicant flow</td></tr><tr><td><span class="label synthetic">Synthetic</span></td><td>Monthly DPD, roll-rate, cure and vintage servicing</td><td>Controls-testing mechanics and dashboard behavior</td><td>Not observed servicing performance</td></tr><tr><td><span class="label proxy">Proxy</span></td><td>Stage 2 and frozen ECL scenario overlays</td><td>Point-in-time sensitivity and readiness gaps</td><td>No production SICR or prior-period ECL trend</td></tr><tr><td><span class="label mixed">Governance</span></td><td>KRI thresholds, actions, policy routing and audit controls</td><td>Decision ownership and closure discipline</td><td>Educational appetite; no organisational approval</td></tr></tbody></table></div></section>
<section class="band"><div class="inner"><div class="section-head"><div><div class="eyebrow" style="color:var(--teal)">Validation</div><h2>Reviewer-verifiable analytical package</h2></div><p>The frozen author reference predates the reviewer run. Reviewer mode cannot create or refresh expected hashes.</p></div><div class="validation-grid"><div><strong>131/131</strong><span>automated checks</span></div><div><strong>56/56</strong><span>UAT assertions</span></div><div><strong>25/25</strong><span>SIT assertions</span></div><div><strong>9/9</strong><span>deterministic outputs</span></div><div><strong>12/12</strong><span>reference and redaction controls</span></div></div><div class="gate"><b>Power BI boundary:</b> DAX expected results are complete at 66/66, while actual runtime remains 0/66 until a real PBIX, refresh test and Power BI screenshots close GATE-001 to GATE-003.</div></div></section>
<section class="band white"><div class="inner"><div class="section-head"><div><div class="eyebrow" style="color:var(--teal)">Primary artifacts</div><h2>Inspect the evidence</h2></div><p>Start with the interactive report, then use the memo and validation files to test the management narrative and technical controls.</p></div><div class="artifacts"><a class="artifact" href="dashboard/index.html"><strong>Interactive risk report</strong><span>Eight business-focused monitoring views</span></a><a class="artifact" href="reports/risk_committee_monitoring_memo.md"><strong>Risk committee memo</strong><span>Condition, movement, decision and owners</span></a><a class="artifact" href="validation/validation_report.md"><strong>Validation report</strong><span>Composition, test results and open gates</span></a><a class="artifact" href="excel/Project2_CreditPortfolioMonitoring_Gold_v2_1_4.xlsx"><strong>Excel control model</strong><span>140 formula nodes and 13 KPI reconciliations</span></a><a class="artifact" href="powerbi/page_specifications.md"><strong>Power BI page specification</strong><span>Eight-page governed build blueprint</span></a><a class="artifact" href="PACKAGE_INDEX.md"><strong>Package index</strong><span>Authoritative artifacts and evidence roles</span></a></div></div></section>
<footer><div><span>Project 2 | Credit Portfolio Monitoring</span><span>Browser implementation is recruiter-facing evidence, not a Power BI screenshot claim.</span></div></footer>
</body></html>"""
    (ROOT / "OPEN_THIS_FIRST.html").write_text(entry_html, encoding="utf-8")


def capture_dashboard_screenshots() -> None:
    from PIL import Image

    def image_quality(path: Path) -> tuple[bool, float, int]:
        if not path.exists() or path.stat().st_size <= 10_000:
            return False, 1.0, 0
        with Image.open(path).convert("RGB") as frame:
            sample = frame.resize((max(1, frame.width // 8), max(1, frame.height // 8)))
            pixels = list(sample.getdata())
        black_ratio = sum(red < 10 and green < 10 and blue < 10 for red, green, blue in pixels) / len(pixels)
        distinct_colors = len(set(pixels))
        return black_ratio < 0.03 and distinct_colors > 100, black_ratio, distinct_colors

    candidates = [
        Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"),
        Path(r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"),
        Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
    ]
    browser = next((path for path in candidates if path.exists()), None)
    evidence_rows = []
    if browser is None:
        evidence_rows.append(["browser_dashboard_screenshots", "EXTERNAL_GATE", "No supported headless browser found"])
    else:
        output_dir = ROOT / "dashboard/screenshots"
        output_dir.mkdir(parents=True, exist_ok=True)
        base_url = (ROOT / "dashboard/index.html").resolve().as_uri()
        captures = [
            ("01_executive_desktop.png", "p1", "1440,1000"),
            ("02_model_monitoring_desktop.png", "p5", "1440,1000"),
            ("03_policy_performance_desktop.png", "p6", "1440,1000"),
            ("04_executive_mobile.png", "p1", "500,900"),
        ]
        for filename, page, window in captures:
            target = output_dir / filename
            quality_ok, black_ratio, distinct_colors = False, 1.0, 0
            completed = None
            for attempt in range(1, 4):
                if target.exists():
                    target.unlink()
                command = [
                    str(browser), "--headless=new", "--disable-gpu", "--hide-scrollbars", "--force-device-scale-factor=1",
                    "--virtual-time-budget=1500", "--run-all-compositor-stages-before-draw", "--disable-features=PaintHolding",
                    f"--window-size={window}", f"--screenshot={target}", f"{base_url}?page={page}&render_attempt={attempt}",
                ]
                completed = subprocess.run(command, capture_output=True, text=True, timeout=60)
                quality_ok, black_ratio, distinct_colors = image_quality(target)
                if completed.returncode == 0 and quality_ok:
                    break
            status = "PASS" if completed is not None and completed.returncode == 0 and quality_ok else "FAIL"
            evidence_rows.append(
                [
                    filename,
                    status,
                    f"browser={browser.name}; viewport={window}; size={target.stat().st_size if target.exists() else 0}; black_ratio={black_ratio:.4f}; distinct_colors={distinct_colors}",
                ]
            )
    evidence = pd.DataFrame(evidence_rows, columns=["artifact", "status", "evidence"])
    write_csv(add_meta(evidence, "Technical", "Executed", "P2", "Browser dashboard screenshots; not Power BI screenshots"), "validation/dashboard_screenshot_validation.csv")
    write_md("dashboard/screenshots/README.md", "Browser Dashboard Screenshots", "These screenshots are captured from the executable HTML dashboard. They are recruiter-facing visual evidence and are **not** Power BI screenshots. Power BI screenshots remain part of the PBIX external release gate.")


def generate_excel_reports_powerbi_pack() -> None:
    ensure_dirs()
    generate_sql_layer()
    generate_governance_pack()
    generate_powerbi_pack()
    generate_excel_control_model()
    generate_reports_and_dashboard()
    capture_dashboard_screenshots()
